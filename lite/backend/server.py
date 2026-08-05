"""Agrocorp Lite — a stripped-down post-sales & site-ops admin.

Roles: admin, accounts, post_sales, site_manager
No email, no cron. In-app notifications only.
"""
from __future__ import annotations

import io
import os
import uuid
import logging
from datetime import datetime, timezone
from typing import List, Literal, Optional

import bcrypt
import jwt
from dotenv import load_dotenv
from fastapi import (
    Depends, FastAPI, File, Form, HTTPException, Header, UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.routing import APIRouter
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, EmailStr, Field

load_dotenv()
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("agrocorp-lite")

# ------------------------------------------------------------------ config -
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ.get("JWT_SECRET", "change-me-in-prod")
JWT_ALG = "HS256"
JWT_TTL_MIN = int(os.environ.get("JWT_TTL_MIN", "720"))  # 12h
ADMIN_PHONE = os.environ.get("ADMIN_PHONE", "9999999999")
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@agrocorp.local")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="Agrocorp Lite")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"],
)
api = APIRouter(prefix="/api")


# ---------------------------------------------------------------- helpers -
def new_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_pw(pw: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), h.encode())
    except Exception:
        return False


def make_token(user_id: str) -> str:
    now_ts = datetime.now(timezone.utc)
    payload = {"sub": user_id,
               "iat": int(now_ts.timestamp()),
               "exp": int(now_ts.timestamp()) + JWT_TTL_MIN * 60}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


# ----------------------------------------------------------------- models -
Role = Literal["admin", "accounts", "post_sales", "site_manager"]


class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str = Field(default_factory=lambda: new_id("usr"))
    name: str
    phone: str
    email: Optional[EmailStr] = None
    role: Role
    project_id: Optional[str] = None   # site_managers are scoped to one project
    is_active: bool = True
    must_reset_password: bool = True
    created_at: str = Field(default_factory=now)


class UserCreate(BaseModel):
    name: str
    phone: str
    email: Optional[EmailStr] = None
    role: Role
    project_id: Optional[str] = None


class LoginRequest(BaseModel):
    phone: str
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    project_id: str = Field(default_factory=lambda: new_id("proj"))
    name: str
    location: str = ""
    site_manager_id: Optional[str] = None
    created_at: str = Field(default_factory=now)


class ProjectCreate(BaseModel):
    name: str
    location: str = ""
    site_manager_id: Optional[str] = None


class Unit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    unit_id: str = Field(default_factory=lambda: new_id("unit"))
    project_id: str
    plot_number: str
    area_sqft: float = 0
    plc_details: dict = {}          # {east_facing:X, hill_view:Y, corner:Z}
    other_charges: dict = {}        # {infra_dev, legal, club, maintenance, ifms, gst_rate, sheet_grand_total}
    status: Literal["available", "sold"] = "available"
    # sale details (filled by post_sales)
    buyer_name: Optional[str] = None
    buyer_contact: Optional[str] = None
    sale_date: Optional[str] = None
    final_price: float = 0
    booking_amount: float = 0
    sold_by: Optional[str] = None
    sold_at: Optional[str] = None
    created_at: str = Field(default_factory=now)


class ScheduleRow(BaseModel):
    due_date: str
    amount: float
    notes: str = ""


class SellUnitRequest(BaseModel):
    buyer_name: str
    buyer_contact: str
    sale_date: str
    final_price: float
    booking_amount: float
    schedule: List[ScheduleRow]


class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    payment_id: str = Field(default_factory=lambda: new_id("pay"))
    unit_id: str
    project_id: str
    seq: int                         # 1..N in schedule order
    due_date: str
    amount: float
    notes: str = ""
    status: Literal["pending", "received"] = "pending"
    received_date: Optional[str] = None
    received_notes: str = ""
    marked_by: Optional[str] = None
    marked_at: Optional[str] = None


class PaymentUpdate(BaseModel):
    status: Literal["pending", "received"]
    received_date: Optional[str] = None
    received_notes: str = ""


class ProcurementItem(BaseModel):
    name: str
    quantity: float = Field(ge=0)
    unit: str = "pcs"
    est_cost: float = 0
    notes: str = ""


class ProcurementRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    request_id: str = Field(default_factory=lambda: new_id("proc"))
    project_id: str
    subject: str
    items: List[ProcurementItem]
    priority: Literal["low", "medium", "high", "urgent"] = "medium"
    notes: str = ""
    status: Literal["pending_admin", "pending_clarification",
                    "approved", "rejected",
                    "paid"] = "pending_admin"
    requested_by: str
    requested_at: str = Field(default_factory=now)
    admin_action_by: Optional[str] = None
    admin_action_at: Optional[str] = None
    admin_note: str = ""
    # accounts payment update (after PO)
    po_number: Optional[str] = None
    paid_amount: float = 0
    paid_date: Optional[str] = None
    paid_by: Optional[str] = None
    paid_at: Optional[str] = None
    paid_notes: str = ""


class ProcurementCreate(BaseModel):
    project_id: str
    subject: str
    items: List[ProcurementItem]
    priority: Literal["low", "medium", "high", "urgent"] = "medium"
    notes: str = ""


class AdminAction(BaseModel):
    action: Literal["approve", "reject", "clarify"]
    note: str = ""


class ProcurementPayment(BaseModel):
    po_number: str
    paid_amount: float
    paid_date: str
    notes: str = ""


class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    item_id: str = Field(default_factory=lambda: new_id("inv"))
    project_id: str
    name: str
    quantity: float = 0
    unit: str = "pcs"
    notes: str = ""
    updated_by: Optional[str] = None
    updated_at: str = Field(default_factory=now)


class InventoryCreate(BaseModel):
    project_id: str
    name: str
    quantity: float = 0
    unit: str = "pcs"
    notes: str = ""


class InventoryUpdate(BaseModel):
    name: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    notes: Optional[str] = None


class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    notification_id: str = Field(default_factory=lambda: new_id("ntf"))
    user_id: str
    kind: str
    message: str
    link: Optional[str] = None
    is_read: bool = False
    created_at: str = Field(default_factory=now)


# ------------------------------------------------------------------ auth --
async def get_current_user(authorization: Optional[str] = Header(None)) -> User:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing bearer token")
    token = authorization.split(" ", 1)[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.PyJWTError:
        raise HTTPException(401, "Invalid or expired token")
    doc = await db.users.find_one({"user_id": payload["sub"]}, {"_id": 0})
    if not doc or not doc.get("is_active", True):
        raise HTTPException(401, "User not found or inactive")
    return User(**doc)


def require_roles(*roles: Role):
    async def _dep(user: User = Depends(get_current_user)) -> User:
        if user.role not in roles:
            raise HTTPException(403, f"Role {user.role} not permitted")
        return user
    return _dep


async def notify(user_id: str, kind: str, message: str,
                 link: Optional[str] = None) -> None:
    n = Notification(user_id=user_id, kind=kind, message=message, link=link)
    await db.notifications.insert_one(n.model_dump())


async def notify_role(role: Role, kind: str, message: str,
                      link: Optional[str] = None) -> None:
    async for u in db.users.find({"role": role, "is_active": True},
                                  {"_id": 0, "user_id": 1}):
        await notify(u["user_id"], kind, message, link)


# --------------------------------------------------------------- endpoints -
@app.get("/health")
@app.get("/api/health")
async def health():
    return {"status": "ok"}


@api.post("/auth/login")
async def login(payload: LoginRequest):
    doc = await db.users.find_one({"phone": payload.phone}, {"_id": 0})
    if not doc or not doc.get("is_active", True):
        raise HTTPException(401, "Invalid phone or password")
    if not verify_pw(payload.password, doc["password_hash"]):
        raise HTTPException(401, "Invalid phone or password")
    token = make_token(doc["user_id"])
    safe = {k: v for k, v in doc.items() if k != "password_hash"}
    return {"access_token": token, "user": safe}


@api.get("/auth/me")
async def me(user: User = Depends(get_current_user)):
    return user.model_dump()


@api.post("/auth/change-password")
async def change_password(payload: ChangePasswordRequest,
                          user: User = Depends(get_current_user)):
    doc = await db.users.find_one({"user_id": user.user_id})
    if not verify_pw(payload.current_password, doc["password_hash"]):
        raise HTTPException(400, "Current password is incorrect")
    if len(payload.new_password) < 8:
        raise HTTPException(400, "New password must be at least 8 characters")
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"password_hash": hash_pw(payload.new_password),
                  "must_reset_password": False}})
    return {"ok": True}


# ----- users (admin only) -------------------------------------------------
@api.get("/users")
async def list_users(user: User = Depends(require_roles("admin"))):
    users = []
    async for u in db.users.find({}, {"_id": 0, "password_hash": 0}):
        users.append(u)
    return users


@api.post("/users")
async def create_user(payload: UserCreate,
                      user: User = Depends(require_roles("admin"))):
    exists = await db.users.find_one({"phone": payload.phone})
    if exists:
        raise HTTPException(400, "A user with this phone already exists")
    if payload.role == "site_manager" and not payload.project_id:
        raise HTTPException(400, "site_manager must have a project_id")
    u = User(**payload.model_dump())
    doc = u.model_dump()
    doc["password_hash"] = hash_pw(payload.phone)  # initial password = phone
    await db.users.insert_one(doc)
    doc.pop("password_hash", None)
    doc.pop("_id", None)
    return doc


@api.patch("/users/{user_id}")
async def update_user(user_id: str, payload: UserCreate,
                      user: User = Depends(require_roles("admin"))):
    updates = payload.model_dump(exclude_unset=True)
    r = await db.users.update_one({"user_id": user_id}, {"$set": updates})
    if r.matched_count == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}


@api.post("/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str,
                                user: User = Depends(require_roles("admin"))):
    doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "User not found")
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"password_hash": hash_pw(doc["phone"]),
                  "must_reset_password": True}})
    return {"ok": True, "message": f"Password reset to phone number for {doc['name']}"}


@api.delete("/users/{user_id}")
async def delete_user(user_id: str,
                      user: User = Depends(require_roles("admin"))):
    if user_id == user.user_id:
        raise HTTPException(400, "You cannot delete yourself")
    r = await db.users.delete_one({"user_id": user_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}


# ----- projects (admin only) ----------------------------------------------
@api.get("/projects")
async def list_projects(user: User = Depends(get_current_user)):
    q = {}
    if user.role == "site_manager" and user.project_id:
        q = {"project_id": user.project_id}
    return await db.projects.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.post("/projects")
async def create_project(payload: ProjectCreate,
                          user: User = Depends(require_roles("admin"))):
    p = Project(**payload.model_dump())
    await db.projects.insert_one(p.model_dump())
    if payload.site_manager_id:
        await db.users.update_one(
            {"user_id": payload.site_manager_id, "role": "site_manager"},
            {"$set": {"project_id": p.project_id}})
    return p.model_dump()


@api.patch("/projects/{project_id}")
async def update_project(project_id: str, payload: ProjectCreate,
                          user: User = Depends(require_roles("admin"))):
    r = await db.projects.update_one(
        {"project_id": project_id}, {"$set": payload.model_dump(exclude_unset=True)})
    if r.matched_count == 0:
        raise HTTPException(404, "Project not found")
    if payload.site_manager_id:
        await db.users.update_one(
            {"user_id": payload.site_manager_id, "role": "site_manager"},
            {"$set": {"project_id": project_id}})
    return {"ok": True}


@api.delete("/projects/{project_id}")
async def delete_project(project_id: str,
                          user: User = Depends(require_roles("admin"))):
    r = await db.projects.delete_one({"project_id": project_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Project not found")
    await db.units.delete_many({"project_id": project_id})
    return {"ok": True}


# ----- units --------------------------------------------------------------
@api.get("/units")
async def list_units(project_id: Optional[str] = None,
                      status: Optional[str] = None,
                      user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    if status:
        q["status"] = status
    if user.role == "site_manager" and user.project_id:
        q["project_id"] = user.project_id
    return await db.units.find(q, {"_id": 0}).sort("plot_number", 1).to_list(2000)


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


@api.post("/units/import")
async def import_units(project_id: str = Form(...),
                        file: UploadFile = File(...),
                        user: User = Depends(require_roles("admin"))):
    """Import units from a Vacation Village-style RERA cost sheet.

    Columns matched (case-insensitive, whitespace-tolerant):
    UNIT NO. · EXTENT (SFT) · Basic Sale Price · East Facing PLC · Hill View PLC ·
    CORNER PLC · Infrastructure & development charges · GST 18% ·
    Legal and Administrative Charges · GST 18% · Club membership · GST 18% ·
    Advance maintenance Charges for 2 years · GST 18% · IFMS · Grand Total
    """
    proj = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not proj:
        raise HTTPException(404, "Project not found")
    raw = await file.read()
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    # Find header row
    header_idx = None
    for i, row in enumerate(rows[:10]):
        joined = " ".join(str(c or "").lower() for c in row)
        if "unit" in joined and "basic sale price" in joined:
            header_idx = i
            break
    if header_idx is None:
        raise HTTPException(400, "Header row not found in Excel")
    header = [str(c or "").strip() for c in rows[header_idx]]

    def fuzzy(fragment: str) -> Optional[int]:
        for i, h in enumerate(header):
            if fragment.lower() in h.strip().lower():
                return i
        return None

    idx = {
        "unit_no": fuzzy("unit no"),
        "extent": fuzzy("extent"),
        "bsp": fuzzy("basic sale"),
        "east": fuzzy("east facing"),
        "hill": fuzzy("hill view"),
        "corner": fuzzy("corner"),
        "infra": fuzzy("infrastructure"),
        "legal": fuzzy("legal"),
        "club": fuzzy("club"),
        "maint": fuzzy("advance maintenance"),
        "ifms": fuzzy("interest free maintenance") or fuzzy("ifms"),
        "grand": fuzzy("grand total"),
    }

    inserted, errors = 0, []
    for line, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not any(row):
            continue
        try:
            u_no = row[idx["unit_no"]] if idx["unit_no"] is not None else None
            if u_no is None or str(u_no).strip() == "":
                continue
            plot = str(u_no).strip()
            if plot.endswith(".0"):
                plot = plot[:-2]
            area = _num(row[idx["extent"]] if idx["extent"] is not None else 0)
            bsp = _num(row[idx["bsp"]] if idx["bsp"] is not None else 0)
            plc = {
                "east_facing": _num(row[idx["east"]] if idx["east"] is not None else 0),
                "hill_view": _num(row[idx["hill"]] if idx["hill"] is not None else 0),
                "corner": _num(row[idx["corner"]] if idx["corner"] is not None else 0),
            }
            other = {
                "bsp": bsp,
                "infra_dev": _num(row[idx["infra"]] if idx["infra"] is not None else 0),
                "legal": _num(row[idx["legal"]] if idx["legal"] is not None else 0),
                "club": _num(row[idx["club"]] if idx["club"] is not None else 0),
                "maintenance": _num(row[idx["maint"]] if idx["maint"] is not None else 0),
                "ifms": _num(row[idx["ifms"]] if idx["ifms"] is not None else 0),
                "gst_rate": 0.18,
                "sheet_grand_total": _num(row[idx["grand"]] if idx["grand"] is not None else 0),
            }
            existing = await db.units.find_one(
                {"project_id": project_id, "plot_number": plot},
                {"_id": 0, "unit_id": 1, "status": 1})
            payload = {
                "project_id": project_id,
                "plot_number": plot,
                "area_sqft": area,
                "plc_details": plc,
                "other_charges": other,
            }
            if existing:
                if existing.get("status", "available") != "sold":
                    await db.units.update_one({"unit_id": existing["unit_id"]},
                                               {"$set": payload})
                else:
                    errors.append({"row": line, "plot": plot,
                                   "error": "already sold; not overwritten"})
                    continue
            else:
                await db.units.insert_one(Unit(**payload).model_dump())
            inserted += 1
        except Exception as e:
            errors.append({"row": line, "error": str(e)})
    return {"inserted": inserted, "errors": errors}


@api.post("/units/{unit_id}/sell")
async def sell_unit(unit_id: str, payload: SellUnitRequest,
                     user: User = Depends(require_roles("post_sales", "admin"))):
    unit = await db.units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(404, "Unit not found")
    if unit["status"] == "sold":
        raise HTTPException(400, "Unit is already sold")
    if not payload.schedule:
        raise HTTPException(400, "Payment schedule cannot be empty")
    scheduled_total = sum(r.amount for r in payload.schedule)
    remainder = payload.final_price - payload.booking_amount
    if abs(scheduled_total - remainder) > 1:
        raise HTTPException(
            400,
            f"Schedule total ({scheduled_total:,.0f}) must equal "
            f"final_price − booking_amount ({remainder:,.0f})")
    # Update unit
    await db.units.update_one(
        {"unit_id": unit_id},
        {"$set": {"status": "sold",
                  "buyer_name": payload.buyer_name,
                  "buyer_contact": payload.buyer_contact,
                  "sale_date": payload.sale_date,
                  "final_price": payload.final_price,
                  "booking_amount": payload.booking_amount,
                  "sold_by": user.user_id,
                  "sold_at": now()}})
    # Create payment records
    docs = []
    for seq, row in enumerate(payload.schedule, start=1):
        docs.append(Payment(
            unit_id=unit_id, project_id=unit["project_id"],
            seq=seq, due_date=row.due_date, amount=row.amount,
            notes=row.notes).model_dump())
    if docs:
        await db.payments.insert_many(docs)
        for d in docs:
            d.pop("_id", None)
    # Notify admin + accounts (in-app only)
    msg = (f"Sale recorded · Plot {unit['plot_number']} · "
           f"{payload.buyer_name} · \u20B9{payload.final_price:,.0f} "
           f"· {len(docs)} installments")
    link = f"/sales"
    await notify_role("admin", "sale_recorded", msg, link)
    await notify_role("accounts", "sale_recorded", msg, link)
    return {"ok": True, "payments": docs}


# ----- payments (accounts) ------------------------------------------------
@api.get("/payments")
async def list_payments(project_id: Optional[str] = None,
                         status: Optional[str] = None,
                         unit_id: Optional[str] = None,
                         user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    if status:
        q["status"] = status
    if unit_id:
        q["unit_id"] = unit_id
    return await db.payments.find(q, {"_id": 0}).sort("due_date", 1).to_list(2000)


@api.patch("/payments/{payment_id}")
async def update_payment(payment_id: str, payload: PaymentUpdate,
                          user: User = Depends(require_roles("accounts", "admin"))):
    updates = {
        "status": payload.status,
        "received_date": payload.received_date,
        "received_notes": payload.received_notes,
        "marked_by": user.user_id,
        "marked_at": now(),
    }
    r = await db.payments.update_one({"payment_id": payment_id}, {"$set": updates})
    if r.matched_count == 0:
        raise HTTPException(404, "Payment not found")
    pay = await db.payments.find_one({"payment_id": payment_id}, {"_id": 0})
    unit = await db.units.find_one({"unit_id": pay["unit_id"]}, {"_id": 0})
    if payload.status == "received":
        await notify_role(
            "admin", "payment_received",
            f"Payment received · Plot {unit.get('plot_number')} · "
            f"\u20B9{pay['amount']:,.0f}", "/sales")
    return {"ok": True}


# ----- procurement (site_manager -> admin -> accounts) -------------------
@api.get("/procurement")
async def list_procurement(user: User = Depends(get_current_user)):
    q: dict = {}
    if user.role == "site_manager" and user.project_id:
        q["project_id"] = user.project_id
    return await db.procurement.find(q, {"_id": 0}).sort("requested_at", -1).to_list(500)


@api.post("/procurement")
async def create_procurement(payload: ProcurementCreate,
                              user: User = Depends(require_roles("site_manager", "admin"))):
    if user.role == "site_manager" and user.project_id != payload.project_id:
        raise HTTPException(403, "Project not in your scope")
    if not payload.items:
        raise HTTPException(400, "At least one item is required")
    r = ProcurementRequest(**payload.model_dump(), requested_by=user.user_id)
    await db.procurement.insert_one(r.model_dump())
    proj = await db.projects.find_one({"project_id": payload.project_id}, {"_id": 0}) or {}
    msg = (f"Procurement request · {payload.subject} · "
           f"{proj.get('name','')} · Priority: {payload.priority}")
    await notify_role("admin", "procurement_new", msg, "/procurement")
    return r.model_dump()


@api.post("/procurement/{request_id}/action")
async def admin_action_procurement(request_id: str, payload: AdminAction,
                                    user: User = Depends(require_roles("admin"))):
    doc = await db.procurement.find_one({"request_id": request_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Request not found")
    if doc["status"] not in ("pending_admin", "pending_clarification"):
        raise HTTPException(400, f"Cannot act on request in status {doc['status']}")
    if payload.action == "approve":
        new_status = "approved"
    elif payload.action == "reject":
        new_status = "rejected"
    else:
        new_status = "pending_clarification"
    await db.procurement.update_one(
        {"request_id": request_id},
        {"$set": {"status": new_status,
                  "admin_action_by": user.user_id,
                  "admin_action_at": now(),
                  "admin_note": payload.note}})
    msg = f"Procurement '{doc['subject']}' — {payload.action}"
    if payload.note:
        msg += f" · {payload.note}"
    await notify(doc["requested_by"], f"procurement_{payload.action}",
                 msg, "/procurement")
    if new_status == "approved":
        await notify_role("accounts", "procurement_approved",
                          f"Approved procurement ready for PO/payment · "
                          f"{doc['subject']}", "/procurement")
    return {"ok": True, "status": new_status}


@api.post("/procurement/{request_id}/payment")
async def procurement_payment(request_id: str, payload: ProcurementPayment,
                               user: User = Depends(require_roles("accounts", "admin"))):
    doc = await db.procurement.find_one({"request_id": request_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Request not found")
    if doc["status"] != "approved":
        raise HTTPException(400, "Only approved requests can be marked paid")
    await db.procurement.update_one(
        {"request_id": request_id},
        {"$set": {"status": "paid",
                  "po_number": payload.po_number,
                  "paid_amount": payload.paid_amount,
                  "paid_date": payload.paid_date,
                  "paid_notes": payload.notes,
                  "paid_by": user.user_id,
                  "paid_at": now()}})
    await notify_role(
        "admin", "procurement_paid",
        f"Payment completed for procurement '{doc['subject']}' · "
        f"PO {payload.po_number} · \u20B9{payload.paid_amount:,.0f}",
        "/procurement")
    await notify(doc["requested_by"], "procurement_paid",
                  f"Payment completed — PO {payload.po_number}",
                  "/procurement")
    return {"ok": True}


# ----- inventory (site_manager) ------------------------------------------
@api.get("/inventory")
async def list_inventory(project_id: Optional[str] = None,
                          user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    if user.role == "site_manager" and user.project_id:
        q["project_id"] = user.project_id
    return await db.inventory.find(q, {"_id": 0}).sort("name", 1).to_list(500)


@api.post("/inventory")
async def create_inventory(payload: InventoryCreate,
                            user: User = Depends(require_roles("site_manager", "admin"))):
    if user.role == "site_manager" and user.project_id != payload.project_id:
        raise HTTPException(403, "Project not in your scope")
    item = InventoryItem(**payload.model_dump(),
                          updated_by=user.user_id, updated_at=now())
    await db.inventory.insert_one(item.model_dump())
    return item.model_dump()


@api.patch("/inventory/{item_id}")
async def update_inventory(item_id: str, payload: InventoryUpdate,
                            user: User = Depends(require_roles("site_manager", "admin"))):
    doc = await db.inventory.find_one({"item_id": item_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Item not found")
    if user.role == "site_manager" and user.project_id != doc["project_id"]:
        raise HTTPException(403, "Project not in your scope")
    updates = payload.model_dump(exclude_unset=True)
    updates["updated_by"] = user.user_id
    updates["updated_at"] = now()
    await db.inventory.update_one({"item_id": item_id}, {"$set": updates})
    return {"ok": True}


@api.delete("/inventory/{item_id}")
async def delete_inventory(item_id: str,
                            user: User = Depends(require_roles("site_manager", "admin"))):
    doc = await db.inventory.find_one({"item_id": item_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Item not found")
    if user.role == "site_manager" and user.project_id != doc["project_id"]:
        raise HTTPException(403, "Project not in your scope")
    await db.inventory.delete_one({"item_id": item_id})
    return {"ok": True}


# ----- notifications -----------------------------------------------------
@api.get("/notifications")
async def list_notifications(user: User = Depends(get_current_user)):
    return await db.notifications.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(50)


@api.post("/notifications/{nid}/read")
async def mark_notification_read(nid: str,
                                  user: User = Depends(get_current_user)):
    await db.notifications.update_one(
        {"notification_id": nid, "user_id": user.user_id},
        {"$set": {"is_read": True}})
    return {"ok": True}


@api.post("/notifications/read-all")
async def mark_all_read(user: User = Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": user.user_id, "is_read": False},
        {"$set": {"is_read": True}})
    return {"ok": True}


# ----- dashboard ---------------------------------------------------------
@api.get("/dashboard")
async def dashboard(user: User = Depends(get_current_user)):
    proj_q = {}
    if user.role == "site_manager" and user.project_id:
        proj_q = {"project_id": user.project_id}
    projects = await db.projects.count_documents(proj_q)
    units_available = await db.units.count_documents({**proj_q, "status": "available"})
    units_sold = await db.units.count_documents({**proj_q, "status": "sold"})
    payments_pending = await db.payments.count_documents({**proj_q, "status": "pending"})
    procurement_pending = await db.procurement.count_documents(
        {**proj_q, "status": {"$in": ["pending_admin", "pending_clarification"]}})
    return {
        "projects": projects,
        "units_available": units_available,
        "units_sold": units_sold,
        "payments_pending": payments_pending,
        "procurement_pending": procurement_pending,
    }


# ----- startup ------------------------------------------------------------
app.include_router(api)


@app.on_event("startup")
async def startup():
    await db.users.create_index("phone", unique=True)
    # Provision the initial admin if the users collection is empty
    n = await db.users.count_documents({})
    if n == 0:
        admin = User(
            name="Admin", phone=ADMIN_PHONE, email=ADMIN_EMAIL, role="admin",
            must_reset_password=True,
        )
        doc = admin.model_dump()
        doc["password_hash"] = hash_pw(ADMIN_PHONE)
        await db.users.insert_one(doc)
        log.info(
            "Provisioned initial admin: phone=%s (initial password = phone)",
            ADMIN_PHONE)


@app.on_event("shutdown")
async def shutdown():
    client.close()
