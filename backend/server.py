"""
Real Estate Stakeholder Dashboard — FastAPI backend.

Modules: auth (password + JWT + bcrypt), users, projects, unit_types, units,
payments, expenses (2-stage approval), stock (items + movements), audit log,
settings, excel import/export, dashboard analytics, notifications, files,
revenue targets, onboarding.
"""
from __future__ import annotations

import io
import os
import re
import csv
import uuid
import secrets
import logging
import smtplib
from email.message import EmailMessage
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Literal, Any

import jwt
import bcrypt
import requests
from dotenv import load_dotenv
from fastapi import (
    FastAPI, APIRouter, HTTPException, Depends, Header, Query, Response,
    UploadFile, File, Form, Cookie, Request,
)
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------- setup ----
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALG = "HS256"
ACCESS_TOKEN_TTL = timedelta(days=7)
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "deb@agrocorp.co.in").lower()
ADMIN_TEMP_PASSWORD = os.environ.get("ADMIN_TEMP_PASSWORD", "Admin@Agro@2026#")
ADMIN_NAME = os.environ.get("ADMIN_NAME", "Agrocorp Admin")
APP_PUBLIC_URL = os.environ.get("APP_PUBLIC_URL", "")
APP_NAME = "realestate-dashboard"

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="Real Estate Dashboard")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
log = logging.getLogger("app")

# --------------------------------------------------------- object storage ----
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
_storage_key: Optional[str] = None


def init_storage() -> Optional[str]:
    global _storage_key
    if _storage_key:
        return _storage_key
    if not EMERGENT_KEY:
        log.warning("EMERGENT_LLM_KEY missing — object storage disabled")
        return None
    try:
        r = requests.post(f"{STORAGE_URL}/init",
                          json={"emergent_key": EMERGENT_KEY}, timeout=30)
        r.raise_for_status()
        _storage_key = r.json()["storage_key"]
        log.info("Object storage initialized")
        return _storage_key
    except Exception as e:
        log.error("Storage init failed: %s", e)
        return None


def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    if not key:
        raise HTTPException(500, "Storage unavailable")
    r = requests.put(f"{STORAGE_URL}/objects/{path}",
                     headers={"X-Storage-Key": key, "Content-Type": content_type},
                     data=data, timeout=120)
    r.raise_for_status()
    return r.json()


def get_object(path: str) -> tuple[bytes, str]:
    key = init_storage()
    if not key:
        raise HTTPException(500, "Storage unavailable")
    r = requests.get(f"{STORAGE_URL}/objects/{path}",
                     headers={"X-Storage-Key": key}, timeout=60)
    r.raise_for_status()
    return r.content, r.headers.get("Content-Type", "application/octet-stream")


# -------------------------------------------------------------- models -----
Role = Literal["super_admin", "process_admin",
               "sales_head", "sales_rep",
               "accounts_head", "accounts_rep",
               "crm_head", "post_sales_rep",
               "site_supervisor"]
ROLE_ORDER = ["super_admin", "process_admin",
              "crm_head", "sales_head", "accounts_head",
              "sales_rep", "post_sales_rep", "accounts_rep",
              "site_supervisor"]
ROLE_LABELS = {
    "super_admin": "Super Admin",
    "process_admin": "Process Admin",
    "crm_head": "CRM Head",
    "sales_head": "Sales Head",
    "accounts_head": "Accounts Head",
    "sales_rep": "Sales Representative",
    "post_sales_rep": "Post-Sales Representative",
    "accounts_rep": "Accounts Representative",
    "site_supervisor": "Site Supervisor",
}
# Convenience sets used by RBAC checks
ADMIN_TIER = ("super_admin", "process_admin")
HEADS = ("super_admin", "process_admin", "crm_head", "sales_head", "accounts_head")
SETUP_ROLES = ("super_admin", "process_admin")
FINAL_APPROVERS = ("super_admin",)  # only Super Admin gives final approval
ExpenseStatus = Literal["pending", "stage1_approved", "final_approved",
                        "rejected"]


def new_id(prefix: str = "id") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:16]}"


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str = Field(default_factory=lambda: new_id("user"))
    email: EmailStr
    name: str
    phone: Optional[str] = None
    picture: Optional[str] = None
    role: Role = "site_supervisor"
    project_ids: List[str] = []
    password_hash: Optional[str] = None
    must_reset_password: bool = True
    dashboard_config: Optional[dict] = None
    is_active: bool = True
    onboarding_completed: bool = False
    created_at: str = Field(default_factory=now)


class UserCreate(BaseModel):
    email: EmailStr
    name: str
    role: Role
    phone: Optional[str] = None
    project_ids: List[str] = []


class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[Role] = None
    phone: Optional[str] = None
    project_ids: Optional[List[str]] = None
    is_active: Optional[bool] = None


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class DashboardConfig(BaseModel):
    widgets: List[str]  # ordered widget ids the user wants to see


ProjectType = Literal["residential", "plots_land"]


class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    project_id: str = Field(default_factory=lambda: new_id("proj"))
    name: str
    project_type: ProjectType = "residential"
    location: str = ""
    address: str = ""
    city: str = ""
    state: str = ""
    pincode: str = ""
    developer: str = ""
    rera_number: str = ""
    start_date: Optional[str] = None
    expected_completion: Optional[str] = None
    total_units_planned: int = 0
    site_manager_id: Optional[str] = None
    image_url: Optional[str] = None
    is_active: bool = True
    created_at: str = Field(default_factory=now)


class ProjectCreate(BaseModel):
    name: str
    project_type: ProjectType = "residential"
    location: str = ""
    address: str = ""
    city: str = ""
    state: str = ""
    pincode: str = ""
    developer: str = ""
    rera_number: str = ""
    start_date: Optional[str] = None
    expected_completion: Optional[str] = None
    total_units_planned: int = 0
    site_manager_id: Optional[str] = None
    image_url: Optional[str] = None


class PLCEntry(BaseModel):
    label: str
    amount: float = 0


UnitStatus = Literal[
    "available", "on_hold", "temporarily_blocked", "booking_in_progress",
    "booked_pending_sales_approval", "sale_confirmed", "post_sales_active",
    "fully_paid", "registration_pending", "registered",
    "possession_pending", "possession_completed",
    "cancelled", "available_for_resale",
]


class Unit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    unit_id: str = Field(default_factory=lambda: new_id("unit"))
    project_id: str
    plot_number: str  # e.g. P-101, A-12
    size: str = ""  # e.g. "1200 sqft" or "30x40 ft"
    facing: str = ""  # North / East etc
    plcs: List[PLCEntry] = []
    price: float = 0
    status: UnitStatus = "available"
    # Sale details (set by Sales role)
    owner_name: Optional[str] = None
    owner_contact: Optional[str] = None
    owner_email: Optional[str] = None
    discount: float = 0
    total_price: float = 0  # price after PLCs and discount
    payment_plan_template_id: Optional[str] = None
    sold_by: Optional[str] = None
    sold_at: Optional[str] = None
    # Sales Head approval (2-step)
    sales_approved_by: Optional[str] = None
    sales_approved_at: Optional[str] = None
    sales_review_note: Optional[str] = None
    # CRM scheduling
    schedule_created_by: Optional[str] = None
    schedule_created_at: Optional[str] = None
    created_at: str = Field(default_factory=now)


class UnitCreate(BaseModel):
    project_id: str
    plot_number: str
    size: str = ""
    facing: str = ""
    plcs: List[PLCEntry] = []
    price: float = 0


class UnitUpdate(BaseModel):
    plot_number: Optional[str] = None
    size: Optional[str] = None
    facing: Optional[str] = None
    plcs: Optional[List[PLCEntry]] = None
    price: Optional[float] = None


class SellUnitRequest(BaseModel):
    owner_name: str
    owner_contact: str = ""
    owner_email: Optional[str] = None
    discount: float = 0
    total_price: float
    payment_plan_template_id: Optional[str] = None


class PlanStage(BaseModel):
    name: str
    percent: float
    days_from_start: int = 0


class PaymentPlanTemplate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    template_id: str = Field(default_factory=lambda: new_id("tpl"))
    name: str
    description: str = ""
    stages: List[PlanStage]
    created_by: Optional[str] = None
    created_at: str = Field(default_factory=now)


class PaymentPlanTemplateCreate(BaseModel):
    name: str
    description: str = ""
    stages: List[PlanStage]


InstallmentStatus = Literal[
    "upcoming", "due_soon", "due_today", "overdue",
    "promise_to_pay", "payment_claimed",
    "not_reflected", "partial",
    "pending_head_approval", "paid", "rejected", "waived", "rescheduled",
]


class Installment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    installment_id: str = Field(default_factory=lambda: new_id("inst"))
    unit_id: str
    project_id: str
    stage_name: str
    percent: float
    amount: float
    due_date: str  # ISO date — original contractual due date (NEVER overwritten)
    revised_due_date: Optional[str] = None  # If rescheduled
    status: InstallmentStatus = "upcoming"
    # Promise-to-Pay
    promise_amount: float = 0
    promise_date: Optional[str] = None
    promise_notes: str = ""
    # Payment claim (by Post-Sales)
    claimed_amount: float = 0
    claimed_at: Optional[str] = None
    claimed_by: Optional[str] = None
    claim_reference: str = ""
    claim_mode: str = ""
    # Bank verification (by Accounts Rep)
    verified_at: Optional[str] = None
    verified_by: Optional[str] = None
    received_amount: float = 0  # amount actually reflected in bank
    # Final approval (by Accounts Head)
    approved_at: Optional[str] = None
    approved_by: Optional[str] = None
    # Legacy fields (kept for compat)
    initiated_at: Optional[str] = None
    initiated_by: Optional[str] = None
    reflected_at: Optional[str] = None
    reflected_by: Optional[str] = None
    notes: str = ""


class InstallmentCreate(BaseModel):
    stage_name: str
    percent: float
    amount: float
    due_date: str
    notes: str = ""


TicketStatus = Literal["open", "in_progress", "resolved", "closed"]
TicketSeverity = Literal["low", "medium", "high", "critical"]


class Ticket(BaseModel):
    model_config = ConfigDict(extra="ignore")
    ticket_id: str = Field(default_factory=lambda: new_id("tkt"))
    project_id: str
    stock_item_id: Optional[str] = None
    subject: str
    description: str = ""
    severity: TicketSeverity = "medium"
    status: TicketStatus = "open"
    raised_by: str
    resolved_by: Optional[str] = None
    resolution_note: str = ""
    resolved_at: Optional[str] = None
    created_at: str = Field(default_factory=now)


class TicketCreate(BaseModel):
    project_id: str
    stock_item_id: Optional[str] = None
    subject: str
    description: str = ""
    severity: TicketSeverity = "medium"


class TicketResolve(BaseModel):
    status: TicketStatus = "resolved"
    resolution_note: str = ""


class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    payment_id: str = Field(default_factory=lambda: new_id("pay"))
    project_id: str
    unit_id: str
    amount: float
    mode: Literal["cash", "cheque", "bank_transfer", "upi", "card", "other"] = "bank_transfer"
    reference: str = ""
    paid_on: str
    recorded_by: str
    created_at: str = Field(default_factory=now)


class PaymentCreate(BaseModel):
    unit_id: str
    amount: float
    mode: Literal["cash", "cheque", "bank_transfer", "upi", "card", "other"] = "bank_transfer"
    reference: str = ""
    paid_on: str


class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    expense_id: str = Field(default_factory=lambda: new_id("exp"))
    project_id: str
    category: str
    amount: float
    vendor: str = ""
    description: str = ""
    receipt_file_id: Optional[str] = None
    status: ExpenseStatus = "pending"
    raised_by: str
    stage1_by: Optional[str] = None
    stage1_at: Optional[str] = None
    final_by: Optional[str] = None
    final_at: Optional[str] = None
    rejection_reason: Optional[str] = None
    rejected_by: Optional[str] = None
    rejected_at: Optional[str] = None
    created_at: str = Field(default_factory=now)


class ExpenseCreate(BaseModel):
    project_id: str
    category: str
    amount: float
    vendor: str = ""
    description: str = ""
    receipt_file_id: Optional[str] = None


class ApprovalAction(BaseModel):
    action: Literal["approve", "reject"]
    reason: Optional[str] = None


class StockItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    item_id: str = Field(default_factory=lambda: new_id("item"))
    project_id: str
    name: str
    unit: str = "pcs"   # kg, bag, m3, pcs
    opening: float = 0
    inward: float = 0
    outward: float = 0
    vendor: str = ""
    created_at: str = Field(default_factory=now)


class StockItemCreate(BaseModel):
    project_id: str
    name: str
    unit: str = "pcs"
    opening: float = 0
    vendor: str = ""


class StockMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    movement_id: str = Field(default_factory=lambda: new_id("mov"))
    item_id: str
    project_id: str
    kind: Literal["inward", "outward"]
    quantity: float
    note: str = ""
    recorded_by: str
    recorded_at: str = Field(default_factory=now)


class StockMovementCreate(BaseModel):
    item_id: str
    kind: Literal["inward", "outward"]
    quantity: float
    note: str = ""


class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    approval_threshold: float = 50000
    currency: str = "INR"
    company_name: str = "Agrocorp Admin"
    logo_file_id: Optional[str] = None
    updated_at: str = Field(default_factory=now)


class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    notification_id: str = Field(default_factory=lambda: new_id("ntf"))
    user_id: str
    kind: str
    message: str
    link: Optional[str] = None
    is_read: bool = False
    created_at: str = Field(default_factory=now)


PeriodType = Literal["monthly", "quarterly"]


class RevenueTarget(BaseModel):
    model_config = ConfigDict(extra="ignore")
    target_id: str = Field(default_factory=lambda: new_id("tgt"))
    project_id: str
    period_type: PeriodType
    period_key: str   # "YYYY-MM" or "YYYY-Qn"
    amount: float
    created_by: Optional[str] = None
    created_at: str = Field(default_factory=now)


_MONTHLY_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")
_QUARTERLY_RE = re.compile(r"^\d{4}-Q[1-4]$")


class RevenueTargetCreate(BaseModel):
    project_id: str
    period_type: PeriodType
    period_key: str
    amount: float = Field(ge=0)

    def validate_period_key(self):
        pat = _MONTHLY_RE if self.period_type == "monthly" else _QUARTERLY_RE
        if not pat.match(self.period_key):
            raise HTTPException(
                400,
                f"period_key must match {pat.pattern} for {self.period_type}")


def period_key_of(iso_dt: str, kind: PeriodType) -> str:
    d = datetime.fromisoformat(iso_dt.replace("Z", "+00:00"))
    if kind == "monthly":
        return d.strftime("%Y-%m")
    q = (d.month - 1) // 3 + 1
    return f"{d.year}-Q{q}"


def current_period_keys() -> dict[str, str]:
    today = datetime.now(timezone.utc)
    q = (today.month - 1) // 3 + 1
    return {"monthly": today.strftime("%Y-%m"),
            "quarterly": f"{today.year}-Q{q}"}


def prior_period_keys(kind: PeriodType, count: int) -> list[str]:
    """Return list of most recent `count` period keys (oldest first, including current)."""
    today = datetime.now(timezone.utc).replace(day=1)
    keys: list[str] = []
    if kind == "monthly":
        d = today
        for _ in range(count):
            keys.append(d.strftime("%Y-%m"))
            # step back one month
            d = (d - timedelta(days=1)).replace(day=1)
    else:
        y = today.year
        q = (today.month - 1) // 3 + 1
        for _ in range(count):
            keys.append(f"{y}-Q{q}")
            q -= 1
            if q == 0:
                q = 4
                y -= 1
    return list(reversed(keys))


# ---------------------------------------------------------------- helpers ---
async def audit(actor_id: str, action: str, entity: str,
                entity_id: str, meta: dict[str, Any] | None = None):
    await db.audit_logs.insert_one({
        "log_id": new_id("log"),
        "actor_id": actor_id, "action": action, "entity": entity,
        "entity_id": entity_id, "meta": meta or {}, "created_at": now(),
    })


async def notify(user_id: str, kind: str, message: str,
                 link: str | None = None):
    await db.notifications.insert_one(Notification(
        user_id=user_id, kind=kind, message=message, link=link,
    ).model_dump())


async def get_settings_doc() -> Settings:
    doc = await db.settings.find_one({"_id": "singleton"})
    if not doc:
        s = Settings()
        d = s.model_dump()
        d["_id"] = "singleton"
        await db.settings.insert_one(d)
        return s
    doc.pop("_id", None)
    return Settings(**doc)


# ------------------------------------------------------------------ auth ---
def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_pw(pw: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), h.encode())
    except Exception:
        return False


def create_jwt(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id, "email": email,
        "exp": datetime.now(timezone.utc) + ACCESS_TOKEN_TTL,
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


def gen_temp_password(length: int = 12) -> str:
    """Generate a secure temp password: 3 upper + 3 lower + 3 digit + 3 symbol."""
    upper = "ABCDEFGHJKLMNPQRSTUVWXYZ"
    lower = "abcdefghijkmnpqrstuvwxyz"
    digit = "23456789"
    symbol = "@#$%&*!"
    parts = ([secrets.choice(upper) for _ in range(3)] +
             [secrets.choice(lower) for _ in range(3)] +
             [secrets.choice(digit) for _ in range(3)] +
             [secrets.choice(symbol) for _ in range(3)])
    secrets.SystemRandom().shuffle(parts)
    return "".join(parts)


_PW_RE = re.compile(r"^(?=.*[A-Z])(?=.*[a-z])(?=.*\d)(?=.*[^\w\s]).{10,}$")


def validate_password_strength(pw: str) -> None:
    if not _PW_RE.match(pw):
        raise HTTPException(
            400,
            "Password must be at least 10 characters and include upper case, "
            "lower case, a digit and a symbol.",
        )


async def get_current_user(
    request: Request,
    authorization: Optional[str] = Header(None),
    access_token: Optional[str] = Cookie(None),
) -> User:
    token = access_token
    if not token and authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ", 1)[1]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Session expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

    user_doc = await db.users.find_one({"user_id": payload["sub"]}, {"_id": 0})
    if not user_doc or not user_doc.get("is_active", True):
        raise HTTPException(401, "User not found or inactive")
    return User(**user_doc)


def require_roles(*roles: Role):
    async def _dep(user: User = Depends(get_current_user)) -> User:
        if user.role not in roles:
            raise HTTPException(403, f"Requires role in {roles}")
        return user
    return _dep


def user_scope_projects(user: User) -> Optional[List[str]]:
    """Return None (all projects) or list of project_ids the user can see."""
    # Admin tier + heads + sales/CRM reps see all projects (they act across the org).
    global_roles = set(ADMIN_TIER) | set(HEADS) | {"sales_rep", "post_sales_rep", "accounts_rep"}
    if user.role in global_roles:
        return None
    # Site Supervisor is scoped to assigned projects only.
    return user.project_ids or []


def is_setup_role(user: User) -> bool:
    """Users who can create/edit projects, units, and templates (subject to Super Admin approval)."""
    return user.role in SETUP_ROLES


def _mgmt_cannot_touch_admin(actor: User, target: dict) -> None:
    """Process Admin is not allowed to modify Super Admin accounts."""
    if actor.role == "process_admin" and target.get("role") == "super_admin":
        raise HTTPException(403, "Process Admin cannot modify Super Admin accounts")


# --------------------------------------- project-type inventory schemas ----
TYPE_SCHEMAS: dict[str, dict[str, Any]] = {
    "residential": {
        "label": "Residential",
        "description": "Apartments, blocks, towers.",
        "fields": [
            {"key": "plot_number", "label": "Plot / Unit number", "type": "text"},
            {"key": "size", "label": "Size", "type": "text"},
            {"key": "facing", "label": "Facing", "type": "select",
             "options": ["North","East","South","West","North-East","North-West","South-East","South-West"]},
            {"key": "price", "label": "Base price", "type": "number"},
        ],
    },
    "plots_land": {
        "label": "Plots / Land",
        "description": "Freehold residential or commercial plots.",
        "fields": [
            {"key": "plot_number", "label": "Plot number", "type": "text"},
            {"key": "size", "label": "Size / Dimensions", "type": "text"},
            {"key": "facing", "label": "Facing", "type": "select",
             "options": ["North","East","South","West","North-East","North-West","South-East","South-West"]},
            {"key": "price", "label": "Base price", "type": "number"},
        ],
    },
}


# ---------------------------------------------------- email (SMTP) --------
def _smtp_configured() -> bool:
    return bool(os.environ.get("SMTP_HOST")
                and os.environ.get("SMTP_USER")
                and os.environ.get("SMTP_PASSWORD"))


def send_email(to_email: str, subject: str, html_body: str,
               text_body: Optional[str] = None) -> bool:
    """Best-effort SMTP send. Returns True on success, False otherwise."""
    if not _smtp_configured():
        return False
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        from_name = os.environ.get("SMTP_FROM_NAME", "Vista Estates")
        msg["From"] = f'{from_name} <{os.environ["SMTP_USER"]}>'
        msg["To"] = to_email
        msg.set_content(text_body or re.sub(r"<[^>]+>", "", html_body))
        msg.add_alternative(html_body, subtype="html")
        with smtplib.SMTP(os.environ["SMTP_HOST"],
                          int(os.environ.get("SMTP_PORT", 587)),
                          timeout=15) as s:
            s.starttls()
            s.login(os.environ["SMTP_USER"], os.environ["SMTP_PASSWORD"])
            s.send_message(msg)
        return True
    except Exception as e:
        log.warning("SMTP send failed: %s", e)
        return False


def invite_email_html(name: str, email: str, temp_pw: str,
                      login_url: str) -> str:
    return f"""
    <div style="font-family:-apple-system,Segoe UI,Roboto,sans-serif;max-width:560px;margin:auto;color:#1c1917">
      <h2 style="color:#064e3b;margin:0 0 12px">You've been invited to Vista Estates</h2>
      <p>Hi {name},</p>
      <p>An account has been created for you on the Vista Estates Stakeholder
      Console. Use the credentials below to sign in — you'll be asked to set a
      new password of your choosing on first login.</p>
      <div style="background:#f5f5f4;border:1px solid #e7e5e4;border-radius:8px;
                  padding:16px;margin:20px 0;font-family:ui-monospace,Menlo,monospace">
        <div>Portal: <a href="{login_url}" style="color:#064e3b">{login_url}</a></div>
        <div>Login ID: <b>{email}</b></div>
        <div>Temporary password: <b>{temp_pw}</b></div>
      </div>
      <p style="font-size:12px;color:#78716c">This is an automated message.
      If you weren't expecting this, please contact your administrator.</p>
    </div>
    """


# ------------------------------------------------------ auth endpoints -----
@api.post("/auth/login")
async def auth_login(payload: LoginRequest, response: Response):
    email = payload.email.lower()
    key = f"login:{email}"
    now_ts = datetime.now(timezone.utc)
    lock = await db.login_attempts.find_one({"_id": key})

    # If a lockout window is still active, deny with remaining time
    if lock and lock.get("locked_until"):
        lu = datetime.fromisoformat(lock["locked_until"])
        if lu.tzinfo is None:
            lu = lu.replace(tzinfo=timezone.utc)
        if lu > now_ts:
            secs_left = (lu - now_ts).total_seconds()
            mins = min(15, max(1, -(-int(secs_left) // 60)))  # ceil, capped at 15
            raise HTTPException(
                429,
                f"Too many failed attempts. Please try again in "
                f"{mins} minute{'s' if mins != 1 else ''} — or "
                f"use 'Forgot password' to reset instantly.")
        # Cooldown expired — start fresh
        await db.login_attempts.delete_one({"_id": key})
        lock = None

    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not user.get("password_hash") \
            or not verify_pw(payload.password, user["password_hash"]):
        # Increment failed attempts; lock for 15 min on the 5th failure
        prev_count = lock.get("count", 0) if lock else 0
        new_count = prev_count + 1
        locked_until = (now_ts + timedelta(minutes=15)).isoformat() \
            if new_count >= 5 else None
        await db.login_attempts.update_one(
            {"_id": key},
            {"$set": {"count": new_count, "last_at": now(),
                      "locked_until": locked_until}},
            upsert=True,
        )
        raise HTTPException(401, "Invalid email or password")

    if not user.get("is_active", True):
        raise HTTPException(403, "Account deactivated")

    await db.login_attempts.delete_one({"_id": key})
    token = create_jwt(user["user_id"], user["email"])
    response.set_cookie(
        "access_token", token, max_age=int(ACCESS_TOKEN_TTL.total_seconds()),
        httponly=True, secure=True, samesite="none", path="/",
    )
    return {
        "access_token": token,
        "user": {k: v for k, v in user.items() if k != "password_hash"},
        "must_reset_password": user.get("must_reset_password", False),
    }


@api.post("/auth/logout")
async def auth_logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def auth_me(user: User = Depends(get_current_user)):
    return user.model_dump()


@api.post("/auth/change-password")
async def change_password(payload: ChangePasswordRequest,
                          user: User = Depends(get_current_user)):
    doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    if not doc or not verify_pw(payload.current_password,
                                doc.get("password_hash", "")):
        raise HTTPException(400, "Current password is incorrect")
    validate_password_strength(payload.new_password)
    if verify_pw(payload.new_password, doc["password_hash"]):
        raise HTTPException(400, "New password must differ from current one")
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"password_hash": hash_pw(payload.new_password),
                  "must_reset_password": False}})
    await audit(user.user_id, "change_password", "user", user.user_id, {})
    return {"ok": True}


class ForgotPasswordRequest(BaseModel):
    email: str


@api.post("/auth/forgot-password")
async def forgot_password(payload: ForgotPasswordRequest):
    """
    Public — self-service password reset.
    Always returns success so we don't leak which emails are registered.
    If the user exists AND is active, we rotate their password to a fresh
    temp and email it. They'll be forced to reset it on next login.
    """
    email = (payload.email or "").strip().lower()
    generic = {
        "ok": True,
        "message": ("If an account exists with that email, "
                    "a temporary password has been sent."),
    }
    if not email:
        return generic

    # simple throttle: max 5 requests / 15 minutes per email
    key = f"forgot:{email}"
    now_ts = datetime.now(timezone.utc)
    attempt = await db.login_attempts.find_one({"_id": key}) or {}
    window_start = attempt.get("window_start")
    count = attempt.get("count", 0)
    if window_start:
        try:
            w = datetime.fromisoformat(window_start)
            if (now_ts - w).total_seconds() < 900 and count >= 5:
                return generic  # silently drop over-limit
            if (now_ts - w).total_seconds() >= 900:
                count = 0
                window_start = None
        except ValueError:
            pass
    if not window_start:
        window_start = now_ts.isoformat()
    await db.login_attempts.update_one(
        {"_id": key},
        {"$set": {"window_start": window_start, "count": count + 1}},
        upsert=True)

    u = await db.users.find_one({"email": email, "is_active": True},
                                {"_id": 0})
    if not u:
        return generic

    temp_pw = gen_temp_password()
    await db.users.update_one(
        {"user_id": u["user_id"]},
        {"$set": {"password_hash": hash_pw(temp_pw),
                  "must_reset_password": True}})
    # clear failed-login lockout for this account
    await db.login_attempts.delete_one({"_id": f"login:{email}"})
    await audit(u["user_id"], "forgot_password", "user", u["user_id"],
                {"channel": "self_service"})
    login_url = f"{APP_PUBLIC_URL}/login" if APP_PUBLIC_URL else "/login"
    send_email(
        u["email"], "Password reset — your temporary password",
        invite_email_html(u["name"], u["email"], temp_pw, login_url))
    return generic


# ------------------------------------------------------ users --------------
@api.get("/users")
async def list_users(user: User = Depends(require_roles("super_admin", "process_admin"))):
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return docs


@api.post("/users")
async def create_user(payload: UserCreate,
                      user: User = Depends(require_roles("super_admin", "process_admin"))):
    # Management cannot create admins
    if user.role == "process_admin" and payload.role == "super_admin":
        raise HTTPException(403, "Management cannot create Admin accounts")
    existing = await db.users.find_one({"email": payload.email.lower()})
    if existing:
        raise HTTPException(400, "User with this email already exists")
    temp_pw = gen_temp_password()
    u = User(email=payload.email.lower(), name=payload.name,
             role=payload.role, phone=payload.phone,
             project_ids=payload.project_ids,
             password_hash=hash_pw(temp_pw),
             must_reset_password=True)
    doc = u.model_dump()
    await db.users.insert_one(doc)
    await audit(user.user_id, "create_user", "user", u.user_id,
                {"email": u.email, "role": u.role})
    login_url = f"{APP_PUBLIC_URL}/login" if APP_PUBLIC_URL else "/login"
    email_sent = send_email(
        u.email, f"Invite: {os.environ.get('SMTP_FROM_NAME', 'Vista Estates')}",
        invite_email_html(u.name, u.email, temp_pw, login_url))
    doc.pop("password_hash", None)
    doc.pop("_id", None)
    return {
        "user": doc, "temp_password": temp_pw,
        "login_url": login_url, "email_sent": email_sent,
    }


@api.post("/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str,
                               actor: User = Depends(require_roles("super_admin", "process_admin"))):
    """Admin/Management regenerates a temp password for a user."""
    u = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not u:
        raise HTTPException(404, "User not found")
    _mgmt_cannot_touch_admin(actor, u)
    temp_pw = gen_temp_password()
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"password_hash": hash_pw(temp_pw),
                  "must_reset_password": True}})
    await db.login_attempts.delete_one({"_id": f"login:{u['email']}"})
    await audit(actor.user_id, "reset_password", "user", user_id, {})
    login_url = f"{APP_PUBLIC_URL}/login" if APP_PUBLIC_URL else "/login"
    email_sent = send_email(
        u["email"], "Password reset — Vista Estates",
        invite_email_html(u["name"], u["email"], temp_pw, login_url))
    return {"temp_password": temp_pw, "login_url": login_url,
            "email_sent": email_sent}



@api.patch("/users/{user_id}")
async def update_user(user_id: str, payload: UserUpdate,
                      user: User = Depends(require_roles("super_admin", "process_admin"))):
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    _mgmt_cannot_touch_admin(user, target)
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(400, "Nothing to update")
    # Management cannot promote anyone to admin
    if user.role == "process_admin" and update.get("role") == "super_admin":
        raise HTTPException(403, "Management cannot assign Admin role")
    r = await db.users.update_one({"user_id": user_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "User not found")
    await audit(user.user_id, "update_user", "user", user_id, update)
    doc = await db.users.find_one({"user_id": user_id},
                                  {"_id": 0, "password_hash": 0})
    return doc


@api.delete("/users/{user_id}")
async def deactivate_user(user_id: str,
                          user: User = Depends(require_roles("super_admin", "process_admin"))):
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    if user.user_id == user_id:
        raise HTTPException(400, "You cannot deactivate yourself")
    _mgmt_cannot_touch_admin(user, target)
    r = await db.users.update_one({"user_id": user_id},
                                  {"$set": {"is_active": False}})
    if r.matched_count == 0:
        raise HTTPException(404, "User not found")
    await audit(user.user_id, "deactivate_user", "user", user_id, {})
    return {"ok": True}


@api.post("/users/{user_id}/reactivate")
async def reactivate_user(user_id: str,
                          user: User = Depends(require_roles("super_admin", "process_admin"))):
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    _mgmt_cannot_touch_admin(user, target)
    await db.users.update_one({"user_id": user_id},
                              {"$set": {"is_active": True}})
    await audit(user.user_id, "reactivate_user", "user", user_id, {})
    return {"ok": True}


# ------------------------------------------------------ profile (self) ----
@api.patch("/me/profile")
async def update_my_profile(payload: dict,
                            user: User = Depends(get_current_user)):
    """A user updates their own name / phone / picture."""
    allowed = {"name", "phone", "picture"}
    update = {k: v for k, v in payload.items() if k in allowed and v is not None}
    if not update:
        raise HTTPException(400, "Nothing to update")
    await db.users.update_one({"user_id": user.user_id}, {"$set": update})
    await audit(user.user_id, "update_profile", "user", user.user_id, update)
    doc = await db.users.find_one({"user_id": user.user_id},
                                  {"_id": 0, "password_hash": 0})
    return doc


@api.post("/me/picture")
async def upload_my_picture(file: UploadFile = File(...),
                            user: User = Depends(get_current_user)):
    data = await file.read()
    ct = file.content_type or "image/jpeg"
    ext = ct.split("/")[-1] if "/" in ct else "jpg"
    path = f"{APP_NAME}/avatars/{user.user_id}.{ext}"
    result = put_object(path, data, ct)
    file_id = new_id("file")
    from datetime import datetime, timezone
    await db.files.insert_one({
        "file_id": file_id, "storage_path": result["path"],
        "original_filename": file.filename or "avatar",
        "content_type": ct, "size": result.get("size", len(data)),
        "uploaded_by": user.user_id, "is_deleted": False, "is_public": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    pic_url = f"/api/files/{file_id}/download"
    await db.users.update_one({"user_id": user.user_id},
                              {"$set": {"picture": pic_url}})
    return {"picture": pic_url}


# ------------------------------------------------------ projects -----------
@api.get("/projects")
async def list_projects(user: User = Depends(get_current_user)):
    q = {}
    scope = user_scope_projects(user)
    if scope is not None:
        q["project_id"] = {"$in": scope}
    docs = await db.projects.find(q, {"_id": 0}).to_list(1000)
    return docs


@api.post("/projects")
async def create_project(payload: ProjectCreate,
                         user: User = Depends(require_roles("super_admin", "process_admin"))):
    p = Project(**payload.model_dump())
    await db.projects.insert_one(p.model_dump())
    await audit(user.user_id, "create_project", "project", p.project_id,
                payload.model_dump())
    return p.model_dump()


@api.patch("/projects/{project_id}")
async def update_project(project_id: str, payload: ProjectCreate,
                         user: User = Depends(require_roles("super_admin", "process_admin"))):
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    r = await db.projects.update_one({"project_id": project_id}, {"$set": upd})
    if r.matched_count == 0:
        raise HTTPException(404, "Project not found")
    await audit(user.user_id, "update_project", "project", project_id, upd)
    return await db.projects.find_one({"project_id": project_id}, {"_id": 0})


@api.get("/projects/types")
async def project_types_schemas():
    """Public — inventory schema per project type."""
    return TYPE_SCHEMAS


@api.get("/projects/{project_id}/impact")
async def project_impact(project_id: str,
                         user: User = Depends(require_roles("super_admin", "process_admin"))):
    users = await db.users.count_documents({"project_ids": project_id})
    units = await db.units.count_documents({"project_id": project_id})
    payments = await db.payments.count_documents({"project_id": project_id})
    expenses = await db.expenses.count_documents({"project_id": project_id})
    stock = await db.stock_items.count_documents({"project_id": project_id})
    return {"users": users, "units": units, "payments": payments,
            "expenses": expenses, "stock_items": stock}


@api.delete("/projects/{project_id}")
async def delete_project(project_id: str,
                         user: User = Depends(require_roles("super_admin"))):
    n_users = await db.users.count_documents({"project_ids": project_id})
    if n_users > 0:
        raise HTTPException(
            400, f"Cannot delete: {n_users} user(s) still assigned")
    await db.projects.delete_one({"project_id": project_id})
    await db.units.delete_many({"project_id": project_id})
    await db.payments.delete_many({"project_id": project_id})
    await db.expenses.delete_many({"project_id": project_id})
    await db.stock_items.delete_many({"project_id": project_id})
    await db.stock_movements.delete_many({"project_id": project_id})
    await db.installments.delete_many({"project_id": project_id})
    await db.tickets.delete_many({"project_id": project_id})
    await audit(user.user_id, "delete_project", "project", project_id, {})
    return {"ok": True}


# ------------------------------------------------------ units --------------
@api.get("/units")
async def list_units(project_id: Optional[str] = None,
                     user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    scope = user_scope_projects(user)
    if scope is not None:
        if project_id and project_id not in scope:
            return []
        if not project_id:
            q["project_id"] = {"$in": scope}
    docs = await db.units.find(q, {"_id": 0}).to_list(2000)
    return docs


@api.post("/units")
async def create_unit(payload: UnitCreate,
                      user: User = Depends(require_roles("super_admin", "process_admin"))):
    u = Unit(**payload.model_dump())
    await db.units.insert_one(u.model_dump())
    await audit(user.user_id, "create_unit", "unit", u.unit_id,
                {"plot_number": u.plot_number})
    return u.model_dump()


@api.patch("/units/{unit_id}")
async def update_unit(unit_id: str, payload: UnitUpdate,
                      user: User = Depends(require_roles("super_admin", "process_admin"))):
    upd = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if not upd:
        raise HTTPException(400, "Nothing to update")
    r = await db.units.update_one({"unit_id": unit_id}, {"$set": upd})
    if r.matched_count == 0:
        raise HTTPException(404, "Unit not found")
    await audit(user.user_id, "update_unit", "unit", unit_id, upd)
    return await db.units.find_one({"unit_id": unit_id}, {"_id": 0})


@api.post("/units/{unit_id}/sell")
async def mark_sold(unit_id: str, payload: SellUnitRequest,
                    user: User = Depends(require_roles("super_admin", "sales_rep", "sales_head", "process_admin"))):
    unit = await db.units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(404, "Unit not found")
    SOLD_STATES_ANY = ("booked_pending_sales_approval", "sale_confirmed",
                      "post_sales_active", "fully_paid", "registration_pending",
                      "registered", "possession_pending", "possession_completed")
    if unit["status"] in SOLD_STATES_ANY:
        raise HTTPException(400, "Unit is already in the sale pipeline")
    # Sales Rep drafts → pending Sales Head approval. Sales Head / Super Admin
    # entering the sale directly can self-approve (skip the pending state).
    self_approve = user.role in ("sales_head", "super_admin", "process_admin")
    new_status = "sale_confirmed" if self_approve else "booked_pending_sales_approval"
    upd = {
        "status": new_status,
        "owner_name": payload.owner_name,
        "owner_contact": payload.owner_contact,
        "owner_email": payload.owner_email,
        "discount": payload.discount,
        "total_price": payload.total_price,
        "payment_plan_template_id": payload.payment_plan_template_id,
        "sold_by": user.user_id,
        "sold_at": now(),
        "sales_approved_by": user.user_id if self_approve else None,
        "sales_approved_at": now() if self_approve else None,
    }
    await db.units.update_one({"unit_id": unit_id}, {"$set": upd})
    await audit(user.user_id, "sell_unit", "unit", unit_id, upd)
    proj = await db.projects.find_one({"project_id": unit["project_id"]},
                                      {"_id": 0}) or {}
    if self_approve:
        # Directly move to CRM
        msg = (f"Unit {unit.get('plot_number', unit_id)} in "
               f"{proj.get('name', 'project')} sold to {payload.owner_name}")
        async for u in db.users.find(
                {"role": {"$in": ["super_admin", "crm_head", "post_sales_rep", "accounts_head", "accounts_rep"]},
                 "is_active": True}, {"_id": 0}):
            await notify(u["user_id"], "unit_sold", msg,
                         link=f"/crm/{unit_id}")
            send_email(u["email"], "Unit sold — CRM handoff required",
                       f"<p>{msg}</p><p>Total price: "
                       f"₹{payload.total_price:,.0f}</p>")
    else:
        msg = (f"New booking pending your approval: Plot "
               f"{unit.get('plot_number', unit_id)} in "
               f"{proj.get('name', 'project')} — "
               f"₹{payload.total_price:,.0f} · {payload.owner_name}")
        async for u in db.users.find(
                {"role": {"$in": ["sales_head", "super_admin"]}, "is_active": True},
                {"_id": 0}):
            await notify(u["user_id"], "sale_pending_approval", msg,
                         link=f"/sales-approvals/{unit_id}")
            send_email(u["email"], "New booking pending approval",
                       f"<p>{msg}</p>")
    return await db.units.find_one({"unit_id": unit_id}, {"_id": 0})


class SaleReview(BaseModel):
    action: Literal["approve", "reject", "return"]
    note: str = ""


@api.post("/units/{unit_id}/sales-review")
async def sales_head_review(unit_id: str, payload: SaleReview,
                            user: User = Depends(require_roles("super_admin", "sales_head"))):
    unit = await db.units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(404, "Unit not found")
    if unit["status"] != "booked_pending_sales_approval":
        raise HTTPException(400, "Sale is not pending approval")
    if payload.action == "approve":
        await db.units.update_one({"unit_id": unit_id},
                                  {"$set": {"status": "sale_confirmed",
                                            "sales_approved_by": user.user_id,
                                            "sales_approved_at": now(),
                                            "sales_review_note": payload.note}})
        await audit(user.user_id, "approve_sale", "unit", unit_id,
                    {"note": payload.note})
        # notify CRM + Accounts + Post-Sales
        proj = await db.projects.find_one({"project_id": unit["project_id"]},
                                          {"_id": 0}) or {}
        msg = f"Booking approved: Plot {unit['plot_number']} in {proj.get('name','')} · {unit['owner_name']}"
        async for u in db.users.find(
                {"role": {"$in": ["super_admin", "crm_head", "post_sales_rep", "accounts_head", "accounts_rep"]},
                 "is_active": True}, {"_id": 0}):
            await notify(u["user_id"], "sale_approved", msg,
                         link=f"/crm/{unit_id}")
    elif payload.action == "reject":
        await db.units.update_one({"unit_id": unit_id},
                                  {"$set": {"status": "available",
                                            "owner_name": None, "owner_contact": None,
                                            "owner_email": None, "discount": 0,
                                            "total_price": 0,
                                            "payment_plan_template_id": None,
                                            "sales_review_note": payload.note}})
        await audit(user.user_id, "reject_sale", "unit", unit_id,
                    {"note": payload.note})
        # notify original sales rep
        if unit.get("sold_by"):
            await notify(unit["sold_by"], "sale_rejected",
                         f"Your booking for Plot {unit['plot_number']} was rejected. Reason: {payload.note}",
                         link=f"/sales")
    else:  # return
        # Keep pending; sales rep sees the note and re-submits
        await db.units.update_one({"unit_id": unit_id},
                                  {"$set": {"sales_review_note": payload.note}})
        if unit.get("sold_by"):
            await notify(unit["sold_by"], "sale_returned",
                         f"Your booking for Plot {unit['plot_number']} was returned for revision. Note: {payload.note}",
                         link=f"/sales")
    return await db.units.find_one({"unit_id": unit_id}, {"_id": 0})


@api.post("/units/{unit_id}/cancel-sale")
async def cancel_sale(unit_id: str,
                      user: User = Depends(require_roles("super_admin", "process_admin"))):
    unit = await db.units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(404, "Unit not found")
    await db.units.update_one(
        {"unit_id": unit_id},
        {"$set": {"status": "available", "owner_name": None,
                  "owner_contact": None, "owner_email": None,
                  "discount": 0, "total_price": 0,
                  "payment_plan_template_id": None,
                  "sold_by": None, "sold_at": None}})
    await db.installments.delete_many({"unit_id": unit_id})
    await audit(user.user_id, "cancel_sale", "unit", unit_id, {})
    return {"ok": True}


# ------------------------------------- payment plan templates -------------
@api.get("/payment-templates")
async def list_templates(user: User = Depends(get_current_user)):
    return await db.payment_templates.find({}, {"_id": 0}).sort(
        "created_at", -1).to_list(200)


@api.post("/payment-templates")
async def create_template(payload: PaymentPlanTemplateCreate,
                          user: User = Depends(require_roles("super_admin", "process_admin"))):
    total = sum(s.percent for s in payload.stages)
    if abs(total - 100) > 0.01:
        raise HTTPException(400, f"Stages must sum to 100% (got {total})")
    t = PaymentPlanTemplate(**payload.model_dump(), created_by=user.user_id)
    await db.payment_templates.insert_one(t.model_dump())
    await audit(user.user_id, "create_template", "payment_template",
                t.template_id, {"name": t.name})
    return t.model_dump()


@api.delete("/payment-templates/{template_id}")
async def delete_template(template_id: str,
                          user: User = Depends(require_roles("super_admin", "process_admin"))):
    r = await db.payment_templates.delete_one({"template_id": template_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Template not found")
    await audit(user.user_id, "delete_template", "payment_template",
                template_id, {})
    return {"ok": True}


# ---------------------------------------- installments (CRM schedule) ----
@api.get("/units/{unit_id}/installments")
async def list_installments(unit_id: str,
                            user: User = Depends(get_current_user)):
    return await db.installments.find(
        {"unit_id": unit_id}, {"_id": 0}).sort("due_date", 1).to_list(200)


@api.post("/units/{unit_id}/installments")
async def create_installments(unit_id: str, payload: List[InstallmentCreate],
                              user: User = Depends(require_roles("super_admin", "crm_head", "post_sales_rep", "process_admin"))):
    unit = await db.units.find_one({"unit_id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(404, "Unit not found")
    # Replace existing installments
    await db.installments.delete_many({"unit_id": unit_id})
    docs = []
    for p in payload:
        inst = Installment(unit_id=unit_id, project_id=unit["project_id"],
                           stage_name=p.stage_name, percent=p.percent,
                           amount=p.amount, due_date=p.due_date,
                           notes=p.notes)
        docs.append(inst.model_dump())
    if docs:
        await db.installments.insert_many(docs)
    new_status = "post_sales_active" if docs else "sale_confirmed"
    await db.units.update_one(
        {"unit_id": unit_id},
        {"$set": {"status": new_status,
                  "schedule_created_by": user.user_id,
                  "schedule_created_at": now()}})
    await audit(user.user_id, "create_schedule", "unit", unit_id,
                {"count": len(docs)})
    if docs:
        proj = await db.projects.find_one({"project_id": unit["project_id"]},
                                          {"_id": 0}) or {}
        msg = (f"Payment schedule ready for "
               f"{unit.get('plot_number', unit_id)} in "
               f"{proj.get('name', 'project')}")
        async for u in db.users.find(
                {"role": {"$in": ["super_admin", "accounts_head", "accounts_rep"]}, "is_active": True},
                {"_id": 0}):
            await notify(u["user_id"], "schedule_created", msg,
                         link=f"/crm/{unit_id}")
    return {"created": len(docs)}


class PromiseToPayRequest(BaseModel):
    promise_amount: float
    promise_date: str
    notes: str = ""


@api.post("/installments/{installment_id}/promise-to-pay")
async def record_promise_to_pay(installment_id: str, payload: PromiseToPayRequest,
                                user: User = Depends(require_roles("super_admin", "crm_head", "post_sales_rep"))):
    r = await db.installments.update_one(
        {"installment_id": installment_id,
         "status": {"$in": ["upcoming", "due_soon", "due_today", "overdue"]}},
        {"$set": {"status": "promise_to_pay",
                  "promise_amount": payload.promise_amount,
                  "promise_date": payload.promise_date,
                  "promise_notes": payload.notes}})
    if r.matched_count == 0:
        raise HTTPException(404, "Installment not actionable")
    await audit(user.user_id, "promise_to_pay", "installment", installment_id,
                payload.model_dump())
    return await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})


class PaymentClaimRequest(BaseModel):
    claimed_amount: float
    claim_reference: str = ""
    claim_mode: str = "bank_transfer"
    notes: str = ""


@api.post("/installments/{installment_id}/claim")
async def claim_payment(installment_id: str, payload: PaymentClaimRequest,
                        user: User = Depends(require_roles("super_admin", "crm_head", "post_sales_rep"))):
    """Post-Sales records that the customer has initiated a payment."""
    inst = await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})
    if not inst:
        raise HTTPException(404, "Installment not found")
    await db.installments.update_one(
        {"installment_id": installment_id},
        {"$set": {"status": "payment_claimed",
                  "claimed_amount": payload.claimed_amount,
                  "claimed_at": now(),
                  "claimed_by": user.user_id,
                  "claim_reference": payload.claim_reference,
                  "claim_mode": payload.claim_mode,
                  "notes": payload.notes}})
    unit = await db.units.find_one({"unit_id": inst["unit_id"]}, {"_id": 0}) or {}
    async for u in db.users.find(
            {"role": {"$in": ["super_admin", "accounts_head", "accounts_rep"]},
             "is_active": True}, {"_id": 0}):
        await notify(u["user_id"], "payment_claimed",
                     f"₹{payload.claimed_amount:,.0f} claimed for {unit.get('plot_number','')} · verify in bank",
                     link=f"/crm/{inst['unit_id']}")
    await audit(user.user_id, "claim_payment", "installment", installment_id,
                payload.model_dump())
    return await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})


class BankVerifyRequest(BaseModel):
    reflected: bool
    received_amount: float = 0
    reference: str = ""
    notes: str = ""


@api.post("/installments/{installment_id}/verify")
async def verify_bank(installment_id: str, payload: BankVerifyRequest,
                      user: User = Depends(require_roles("super_admin", "accounts_head", "accounts_rep"))):
    """Accounts Rep confirms whether the amount reflected in the bank."""
    inst = await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})
    if not inst:
        raise HTTPException(404, "Installment not found")
    if not payload.reflected:
        await db.installments.update_one(
            {"installment_id": installment_id},
            {"$set": {"status": "not_reflected",
                      "verified_at": now(),
                      "verified_by": user.user_id,
                      "notes": payload.notes}})
        # notify post-sales
        if inst.get("claimed_by"):
            await notify(inst["claimed_by"], "not_reflected",
                         f"Payment not reflected for installment {inst['stage_name']}. Please follow up with customer.",
                         link=f"/crm/{inst['unit_id']}")
        await audit(user.user_id, "verify_not_reflected", "installment",
                    installment_id, payload.model_dump())
        return await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})
    # Reflected — awaits Head approval
    is_partial = payload.received_amount < inst["amount"] - 0.01
    await db.installments.update_one(
        {"installment_id": installment_id},
        {"$set": {"status": "pending_head_approval",
                  "verified_at": now(),
                  "verified_by": user.user_id,
                  "received_amount": payload.received_amount,
                  "claim_reference": payload.reference or inst.get("claim_reference", ""),
                  "notes": payload.notes}})
    async for u in db.users.find(
            {"role": "accounts_head", "is_active": True}, {"_id": 0}):
        await notify(u["user_id"], "verify_pending_approval",
                     f"₹{payload.received_amount:,.0f} verified — please approve" +
                     (" (partial)" if is_partial else ""),
                     link=f"/crm/{inst['unit_id']}")
    await audit(user.user_id, "verify_reflected", "installment",
                installment_id, payload.model_dump())
    return await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})


class AccountsApproveRequest(BaseModel):
    action: Literal["approve", "reject"]
    note: str = ""


@api.post("/installments/{installment_id}/approve")
async def accounts_head_approve(installment_id: str, payload: AccountsApproveRequest,
                                user: User = Depends(require_roles("super_admin", "accounts_head"))):
    """Accounts Head gives final approval, creating a payment record."""
    inst = await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})
    if not inst:
        raise HTTPException(404, "Installment not found")
    if inst["status"] != "pending_head_approval":
        raise HTTPException(400, "Installment not awaiting head approval")
    if payload.action == "reject":
        await db.installments.update_one(
            {"installment_id": installment_id},
            {"$set": {"status": "rejected", "notes": payload.note}})
        await audit(user.user_id, "reject_payment", "installment",
                    installment_id, payload.model_dump())
        return await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})
    # Approve
    is_partial = inst["received_amount"] < inst["amount"] - 0.01
    final_status = "partial" if is_partial else "paid"
    await db.installments.update_one(
        {"installment_id": installment_id},
        {"$set": {"status": final_status,
                  "approved_at": now(),
                  "approved_by": user.user_id}})
    unit = await db.units.find_one({"unit_id": inst["unit_id"]}, {"_id": 0}) or {}
    p = Payment(project_id=unit.get("project_id", ""), unit_id=inst["unit_id"],
                amount=inst["received_amount"], mode=inst.get("claim_mode", "bank_transfer"),
                reference=inst.get("claim_reference", "") or f"Installment {inst['stage_name']}",
                paid_on=now()[:10], recorded_by=user.user_id)
    await db.payments.insert_one(p.model_dump())
    # If all installments paid, mark unit fully_paid
    outstanding = await db.installments.count_documents(
        {"unit_id": inst["unit_id"], "status": {"$nin": ["paid", "waived"]}})
    if outstanding == 0:
        await db.units.update_one({"unit_id": inst["unit_id"]},
                                  {"$set": {"status": "fully_paid"}})
    await audit(user.user_id, "approve_payment", "installment",
                installment_id, {"amount": inst["received_amount"]})
    return await db.installments.find_one({"installment_id": installment_id}, {"_id": 0})


# ---------------------------------------------------------- tickets -------
@api.get("/tickets")
async def list_tickets(project_id: Optional[str] = None,
                       user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    scope = user_scope_projects(user)
    if scope is not None:
        q["project_id"] = q.get("project_id") or {"$in": scope}
    return await db.tickets.find(q, {"_id": 0}).sort(
        "created_at", -1).to_list(500)


@api.post("/tickets")
async def create_ticket(payload: TicketCreate,
                        user: User = Depends(require_roles("site_supervisor", "super_admin", "process_admin", "crm_head"))):
    if user.role == "site_supervisor" and payload.project_id not in user.project_ids:
        raise HTTPException(403, "Project not in your scope")
    t = Ticket(**payload.model_dump(), raised_by=user.user_id)
    await db.tickets.insert_one(t.model_dump())
    await audit(user.user_id, "create_ticket", "ticket", t.ticket_id,
                {"subject": t.subject})
    # notify admin + management
    async for u in db.users.find(
            {"role": {"$in": ["super_admin", "process_admin", "crm_head"]}, "is_active": True},
            {"_id": 0}):
        await notify(u["user_id"], "ticket_new",
                     f"{t.severity.upper()} · {t.subject}",
                     link=f"/tickets/{t.ticket_id}")
    return t.model_dump()


@api.patch("/tickets/{ticket_id}")
async def resolve_ticket(ticket_id: str, payload: TicketResolve,
                         user: User = Depends(require_roles("super_admin", "process_admin"))):
    upd = {"status": payload.status, "resolution_note": payload.resolution_note}
    if payload.status in ("resolved", "closed"):
        upd["resolved_by"] = user.user_id
        upd["resolved_at"] = now()
    r = await db.tickets.update_one({"ticket_id": ticket_id}, {"$set": upd})
    if r.matched_count == 0:
        raise HTTPException(404, "Ticket not found")
    await audit(user.user_id, "resolve_ticket", "ticket", ticket_id, upd)
    return await db.tickets.find_one({"ticket_id": ticket_id}, {"_id": 0})


# ------------------------------------------------------ payments -----------
@api.get("/payments")
async def list_payments(project_id: Optional[str] = None,
                        unit_id: Optional[str] = None,
                        user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    if unit_id:
        q["unit_id"] = unit_id
    scope = user_scope_projects(user)
    if scope is not None:
        if not project_id:
            q["project_id"] = {"$in": scope}
    docs = await db.payments.find(q, {"_id": 0}).sort("paid_on", -1).to_list(2000)
    return docs


@api.post("/payments")
async def create_payment(payload: PaymentCreate,
                         user: User = Depends(require_roles("super_admin", "accounts_head", "accounts_rep"))):
    unit = await db.units.find_one({"unit_id": payload.unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(404, "Unit not found")
    p = Payment(project_id=unit["project_id"],
                unit_id=payload.unit_id, amount=payload.amount,
                mode=payload.mode, reference=payload.reference,
                paid_on=payload.paid_on, recorded_by=user.user_id)
    await db.payments.insert_one(p.model_dump())
    await audit(user.user_id, "record_payment", "payment", p.payment_id,
                {"unit": payload.unit_id, "amount": payload.amount})
    return p.model_dump()


@api.get("/revenue/summary")
async def revenue_summary(project_id: Optional[str] = None,
                          user: User = Depends(get_current_user)):
    scope = user_scope_projects(user)
    unit_q: dict = {}
    if project_id:
        unit_q["project_id"] = project_id
    elif scope is not None:
        unit_q["project_id"] = {"$in": scope}
    units = await db.units.find(unit_q, {"_id": 0}).to_list(5000)
    unit_ids = [u["unit_id"] for u in units]

    pay_q: dict = {"unit_id": {"$in": unit_ids}}
    payments = await db.payments.find(pay_q, {"_id": 0}).to_list(10000)

    SOLD_STATES = ("sale_confirmed", "post_sales_active", "fully_paid",
                   "registration_pending", "registered",
                   "possession_pending", "possession_completed",
                   "booked_pending_sales_approval")
    accrued = sum(u.get("total_price", 0) or u.get("price", 0)
                  for u in units if u["status"] in SOLD_STATES)
    received = sum(p["amount"] for p in payments)
    receivable = accrued - received

    by_unit = {}
    for u in units:
        by_unit[u["unit_id"]] = {
            "unit_id": u["unit_id"],
            "unit_number": u["plot_number"],
            "project_id": u["project_id"],
            "status": u["status"],
            "accrued": (u.get("total_price", 0) or u.get("price", 0))
            if u["status"] in SOLD_STATES else 0,
            "received": 0,
        }
    for p in payments:
        if p["unit_id"] in by_unit:
            by_unit[p["unit_id"]]["received"] += p["amount"]
    for row in by_unit.values():
        row["receivable"] = row["accrued"] - row["received"]

    return {"accrued": accrued, "received": received,
            "receivable": receivable,
            "by_unit": list(by_unit.values())}


# ------------------------------------------------------ revenue targets ---
async def _compute_period_actuals(scope_pids: Optional[list[str]],
                                  period_type: PeriodType,
                                  keys: list[str]) -> dict[str, dict]:
    """
    Return {period_key: {received: X, accrued: Y}} across given projects.
    scope_pids semantics:
      * None      → include ALL projects (no filter)
      * []        → no visible projects → all zeros
      * [ids...]  → limit to these projects
    """
    out: dict[str, dict] = {k: {"received": 0.0, "accrued": 0.0} for k in keys}
    if scope_pids is not None and len(scope_pids) == 0:
        return out
    key_set = set(keys)

    pay_q: dict = {}
    if scope_pids:
        pay_q["project_id"] = {"$in": scope_pids}
    async for p in db.payments.find(pay_q, {"_id": 0}):
        if not p.get("paid_on"):
            continue
        k = period_key_of(p["paid_on"], period_type)
        if k in key_set:
            out[k]["received"] += p["amount"]

    unit_q: dict = {"status": {"$in": ["sale_confirmed", "post_sales_active",
                                       "fully_paid", "registration_pending",
                                       "registered", "possession_pending",
                                       "possession_completed"]}}
    if scope_pids:
        unit_q["project_id"] = {"$in": scope_pids}
    async for u in db.units.find(unit_q, {"_id": 0}):
        if not u.get("sold_at"):
            continue
        k = period_key_of(u["sold_at"], period_type)
        if k in key_set:
            out[k]["accrued"] += u.get("total_price", 0) or u.get("price", 0)
    return out


async def _sum_targets(scope_pids: Optional[list[str]],
                       period_type: PeriodType,
                       keys: list[str]) -> dict[str, float]:
    totals: dict[str, float] = {k: 0.0 for k in keys}
    if scope_pids is not None and len(scope_pids) == 0:
        return totals
    q: dict = {"period_type": period_type, "period_key": {"$in": keys}}
    if scope_pids:
        q["project_id"] = {"$in": scope_pids}
    async for t in db.revenue_targets.find(q, {"_id": 0}):
        totals[t["period_key"]] = totals.get(t["period_key"], 0.0) + t["amount"]
    return totals


def _resolve_scope(user: User, project_id: Optional[str]) -> list[str] | None:
    """Return concrete list of project ids to filter by, or None = no filter."""
    scope = user_scope_projects(user)
    if project_id:
        if scope is not None and project_id not in scope:
            return []
        return [project_id]
    return scope  # may be None


@api.get("/revenue-targets")
async def list_targets(project_id: Optional[str] = None,
                       period_type: Optional[PeriodType] = None,
                       user: User = Depends(get_current_user)):
    q: dict = {}
    scope = _resolve_scope(user, project_id)
    if scope is not None:
        if not scope:
            return []
        q["project_id"] = {"$in": scope}
    if period_type:
        q["period_type"] = period_type
    docs = await db.revenue_targets.find(q, {"_id": 0}).sort("period_key", -1).to_list(500)
    return docs


@api.post("/revenue-targets")
async def upsert_target(payload: RevenueTargetCreate,
                        user: User = Depends(require_roles("super_admin"))):
    payload.validate_period_key()
    if not await db.projects.find_one({"project_id": payload.project_id}):
        raise HTTPException(404, "Project not found")
    existing = await db.revenue_targets.find_one({
        "project_id": payload.project_id,
        "period_type": payload.period_type,
        "period_key": payload.period_key,
    }, {"_id": 0})
    if existing:
        await db.revenue_targets.update_one(
            {"target_id": existing["target_id"]},
            {"$set": {"amount": payload.amount}},
        )
        await audit(user.user_id, "update_target", "revenue_target",
                    existing["target_id"], {"amount": payload.amount})
        return await db.revenue_targets.find_one(
            {"target_id": existing["target_id"]}, {"_id": 0})
    t = RevenueTarget(**payload.model_dump(), created_by=user.user_id)
    await db.revenue_targets.insert_one(t.model_dump())
    await audit(user.user_id, "create_target", "revenue_target", t.target_id,
                payload.model_dump())
    return t.model_dump()


@api.delete("/revenue-targets/{target_id}")
async def delete_target(target_id: str,
                        user: User = Depends(require_roles("super_admin"))):
    r = await db.revenue_targets.delete_one({"target_id": target_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Target not found")
    await audit(user.user_id, "delete_target", "revenue_target", target_id, {})
    return {"ok": True}


@api.get("/revenue-targets/variance")
async def target_variance(project_id: Optional[str] = None,
                          period_type: PeriodType = "monthly",
                          periods: int = 6,
                          user: User = Depends(get_current_user)):
    periods = max(1, min(24, periods))
    scope = _resolve_scope(user, project_id)
    # Preserve None → all-projects; [] → no-access; concrete list → scope
    if scope is None:
        pids = None
    else:
        pids = scope  # may be []
    keys = prior_period_keys(period_type, periods)
    targets = await _sum_targets(pids, period_type, keys)
    actuals = await _compute_period_actuals(pids, period_type, keys)
    series = []
    for k in keys:
        t = targets.get(k, 0.0)
        rec = actuals[k]["received"]
        acc = actuals[k]["accrued"]
        series.append({
            "period_key": k,
            "target": t,
            "received": rec,
            "accrued": acc,
            "variance_received": rec - t,
            "variance_accrued": acc - t,
            "variance_received_pct": None if t == 0 else round(((rec - t) / t) * 100, 1),
            "variance_accrued_pct": None if t == 0 else round(((acc - t) / t) * 100, 1),
        })
    return {"period_type": period_type, "series": series}


# ------------------------------------------------------ expenses -----------
@api.get("/expenses")
async def list_expenses(project_id: Optional[str] = None,
                        status: Optional[ExpenseStatus] = None,
                        user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    if status:
        q["status"] = status
    scope = user_scope_projects(user)
    if scope is not None:
        if not project_id:
            q["project_id"] = {"$in": scope}
    if user.role == "site_supervisor":
        # site managers see only their own + their projects
        q["project_id"] = q.get("project_id", {"$in": user.project_ids})
    docs = await db.expenses.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return docs


@api.post("/expenses")
async def raise_expense(payload: ExpenseCreate,
                        user: User = Depends(require_roles(
                            "site_supervisor", "super_admin", "process_admin"))):
    if user.role == "site_supervisor" and payload.project_id not in user.project_ids:
        raise HTTPException(403, "You are not assigned to this project")
    e = Expense(**payload.model_dump(), raised_by=user.user_id)
    await db.expenses.insert_one(e.model_dump())
    await audit(user.user_id, "raise_expense", "expense", e.expense_id,
                {"amount": payload.amount})
    # notify accounts users
    async for acc in db.users.find({"role": {"$in": ["accounts_rep", "accounts_head"]}, "is_active": True},
                                   {"_id": 0}):
        await notify(acc["user_id"], "expense_new",
                     f"New expense ₹{payload.amount:,.0f} pending Stage-1",
                     link=f"/expenses/{e.expense_id}")
    return e.model_dump()


@api.post("/expenses/{expense_id}/stage1")
async def stage1(expense_id: str, payload: ApprovalAction,
                 user: User = Depends(require_roles("accounts_head", "accounts_rep", "super_admin"))):
    exp = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    if not exp:
        raise HTTPException(404, "Expense not found")
    if exp["status"] != "pending":
        raise HTTPException(400, "Expense not in pending state")

    settings = await get_settings_doc()
    threshold = settings.approval_threshold

    if payload.action == "reject":
        if not payload.reason:
            raise HTTPException(400, "Rejection reason required")
        upd = {"status": "rejected", "rejection_reason": payload.reason,
               "rejected_by": user.user_id, "rejected_at": now()}
    else:
        if exp["amount"] > threshold:
            upd = {"status": "stage1_approved",
                   "stage1_by": user.user_id, "stage1_at": now()}
        else:
            upd = {"status": "final_approved",
                   "stage1_by": user.user_id, "stage1_at": now(),
                   "final_by": user.user_id, "final_at": now()}

    await db.expenses.update_one({"expense_id": expense_id}, {"$set": upd})
    await audit(user.user_id, "stage1", "expense", expense_id, upd)
    if upd["status"] == "stage1_approved":
        async for m in db.users.find({"role": {"$in": ["super_admin", "process_admin", "crm_head"]}, "is_active": True},
                                     {"_id": 0}):
            await notify(m["user_id"], "expense_stage1",
                         f"Expense ₹{exp['amount']:,.0f} needs final approval",
                         link=f"/expenses/{expense_id}")
    else:
        await notify(exp["raised_by"], "expense_result",
                     f"Your expense is {upd['status'].replace('_', ' ')}",
                     link=f"/expenses/{expense_id}")
    return await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})


@api.post("/expenses/{expense_id}/final")
async def final_approve(expense_id: str, payload: ApprovalAction,
                        user: User = Depends(require_roles("super_admin", "process_admin", "crm_head"))):
    exp = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    if not exp:
        raise HTTPException(404, "Expense not found")
    if exp["status"] != "stage1_approved":
        raise HTTPException(400, "Expense not awaiting final approval")

    if payload.action == "reject":
        if not payload.reason:
            raise HTTPException(400, "Rejection reason required")
        upd = {"status": "rejected", "rejection_reason": payload.reason,
               "rejected_by": user.user_id, "rejected_at": now()}
    else:
        upd = {"status": "final_approved",
               "final_by": user.user_id, "final_at": now()}

    await db.expenses.update_one({"expense_id": expense_id}, {"$set": upd})
    await audit(user.user_id, "final", "expense", expense_id, upd)
    await notify(exp["raised_by"], "expense_result",
                 f"Your expense is {upd['status'].replace('_', ' ')}",
                 link=f"/expenses/{expense_id}")
    return await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})


# ------------------------------------------------------ stock --------------
@api.get("/stock/items")
async def list_stock_items(project_id: Optional[str] = None,
                           user: User = Depends(get_current_user)):
    q: dict = {}
    if project_id:
        q["project_id"] = project_id
    scope = user_scope_projects(user)
    if scope is not None:
        if not project_id:
            q["project_id"] = {"$in": scope}
    items = await db.stock_items.find(q, {"_id": 0}).to_list(2000)
    for it in items:
        it["closing"] = (it.get("opening", 0)
                         + it.get("inward", 0) - it.get("outward", 0))
    return items


@api.post("/stock/items")
async def create_stock_item(payload: StockItemCreate,
                            user: User = Depends(require_roles(
                                "super_admin", "process_admin", "site_supervisor"))):
    if user.role == "site_supervisor" and payload.project_id not in user.project_ids:
        raise HTTPException(403, "Project not in your scope")
    it = StockItem(**payload.model_dump())
    await db.stock_items.insert_one(it.model_dump())
    return it.model_dump()


@api.post("/stock/movements")
async def add_movement(payload: StockMovementCreate,
                       user: User = Depends(require_roles(
                           "super_admin", "process_admin", "site_supervisor"))):
    item = await db.stock_items.find_one({"item_id": payload.item_id}, {"_id": 0})
    if not item:
        raise HTTPException(404, "Item not found")
    if (user.role == "site_supervisor"
            and item["project_id"] not in user.project_ids):
        raise HTTPException(403, "Project not in your scope")
    mv = StockMovement(item_id=payload.item_id,
                       project_id=item["project_id"],
                       kind=payload.kind, quantity=payload.quantity,
                       note=payload.note, recorded_by=user.user_id)
    await db.stock_movements.insert_one(mv.model_dump())
    inc = {"inward": payload.quantity} if payload.kind == "inward" \
        else {"outward": payload.quantity}
    await db.stock_items.update_one({"item_id": payload.item_id},
                                    {"$inc": inc})
    await audit(user.user_id, f"stock_{payload.kind}", "stock_item",
                payload.item_id, {"qty": payload.quantity})
    return mv.model_dump()


@api.get("/stock/movements")
async def list_movements(item_id: Optional[str] = None,
                         project_id: Optional[str] = None,
                         user: User = Depends(get_current_user)):
    q: dict = {}
    if item_id:
        q["item_id"] = item_id
    if project_id:
        q["project_id"] = project_id
    return await db.stock_movements.find(q, {"_id": 0}).sort(
        "recorded_at", -1).to_list(2000)


# ------------------------------------------------------ settings -----------
@api.get("/settings/public")
async def public_settings():
    """Non-sensitive settings visible before login (branding)."""
    s = await get_settings_doc()
    return {
        "company_name": s.company_name,
        "currency": s.currency,
        "logo_file_id": s.logo_file_id,
    }


@api.get("/settings")
async def read_settings(user: User = Depends(get_current_user)):
    s = await get_settings_doc()
    return s.model_dump()


class SettingsUpdate(BaseModel):
    approval_threshold: Optional[float] = None
    currency: Optional[str] = None
    company_name: Optional[str] = None
    logo_file_id: Optional[str] = None


@api.patch("/settings")
async def update_settings(payload: SettingsUpdate,
                          user: User = Depends(require_roles("super_admin"))):
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    upd["updated_at"] = now()
    await db.settings.update_one({"_id": "singleton"}, {"$set": upd},
                                 upsert=True)
    await audit(user.user_id, "update_settings", "settings", "singleton", upd)
    return (await get_settings_doc()).model_dump()


# ------------------------------------------------------ notifications -----
@api.get("/notifications")
async def my_notifications(user: User = Depends(get_current_user)):
    return await db.notifications.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(50)


@api.post("/notifications/{nid}/read")
async def mark_read(nid: str, user: User = Depends(get_current_user)):
    await db.notifications.update_one(
        {"notification_id": nid, "user_id": user.user_id},
        {"$set": {"is_read": True}})
    return {"ok": True}


@api.post("/notifications/read-all")
async def mark_all_read(user: User = Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": user.user_id, "is_read": False},
        {"$set": {"is_read": True}})
    return {"ok": True}


# ------------------------------------------------------ audit log ---------
@api.get("/audit-logs")
async def audit_logs(user: User = Depends(require_roles("super_admin"))):
    return await db.audit_logs.find({}, {"_id": 0}).sort(
        "created_at", -1).limit(500).to_list(500)


# ------------------------------------------------------ files -------------
@api.post("/files/upload")
async def upload_file(file: UploadFile = File(...),
                      user: User = Depends(get_current_user)):
    data = await file.read()
    ext = (file.filename or "bin").split(".")[-1].lower()
    file_id = new_id("file")
    path = f"{APP_NAME}/uploads/{user.user_id}/{file_id}.{ext}"
    result = put_object(path, data, file.content_type or "application/octet-stream")
    doc = {
        "file_id": file_id,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size", len(data)),
        "uploaded_by": user.user_id,
        "is_deleted": False,
        "created_at": now(),
    }
    await db.files.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/files/{file_id}/download")
async def download_file(file_id: str,
                        authorization: Optional[str] = Header(None),
                        access_token: Optional[str] = Cookie(None),
                        auth: Optional[str] = Query(None)):
    rec = await db.files.find_one({"file_id": file_id, "is_deleted": False},
                                  {"_id": 0})
    if not rec:
        raise HTTPException(404, "File not found")
    if not rec.get("is_public"):
        token = access_token
        if not token and authorization and authorization.startswith("Bearer "):
            token = authorization.split(" ", 1)[1]
        if not token and auth:
            token = auth
        if not token:
            raise HTTPException(401, "Not authenticated")
        try:
            jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        except Exception:
            raise HTTPException(401, "Invalid token")
    data, ct = get_object(rec["storage_path"])
    return Response(content=data,
                    media_type=rec.get("content_type") or ct,
                    headers={"Content-Disposition":
                             f'inline; filename="{rec["original_filename"]}"'})


@api.post("/files/logo")
async def upload_logo(file: UploadFile = File(...),
                      user: User = Depends(require_roles("super_admin"))):
    data = await file.read()
    ext = (file.filename or "png").split(".")[-1].lower()
    file_id = new_id("file")
    path = f"{APP_NAME}/branding/{file_id}.{ext}"
    result = put_object(path, data, file.content_type or "image/png")
    doc = {
        "file_id": file_id, "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size", len(data)),
        "uploaded_by": user.user_id, "is_deleted": False, "is_public": True,
        "created_at": now(),
    }
    await db.files.insert_one(doc)
    await db.settings.update_one({"_id": "singleton"},
                                 {"$set": {"logo_file_id": file_id,
                                           "updated_at": now()}},
                                 upsert=True)
    await audit(user.user_id, "upload_logo", "settings", "singleton",
                {"file_id": file_id})
    doc.pop("_id", None)
    return doc


@api.post("/projects/{project_id}/image")
async def upload_project_image(project_id: str,
                               file: UploadFile = File(...),
                               user: User = Depends(require_roles("super_admin"))):
    proj = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not proj:
        raise HTTPException(404, "Project not found")
    data = await file.read()
    ext = (file.filename or "jpg").split(".")[-1].lower()
    file_id = new_id("file")
    path = f"{APP_NAME}/projects/{project_id}/{file_id}.{ext}"
    result = put_object(path, data, file.content_type or "image/jpeg")
    doc = {
        "file_id": file_id, "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size", len(data)),
        "uploaded_by": user.user_id, "is_deleted": False, "is_public": True,
        "created_at": now(),
    }
    await db.files.insert_one(doc)
    image_url = f"/api/files/{file_id}/download"
    await db.projects.update_one({"project_id": project_id},
                                 {"$set": {"image_url": image_url}})
    await audit(user.user_id, "upload_project_image", "project", project_id,
                {"file_id": file_id})
    return {"file_id": file_id, "image_url": image_url}



# ------------------------------------------------------ analytics ---------
@api.get("/dashboard/summary")
async def dashboard_summary(project_id: Optional[str] = None,
                            user: User = Depends(get_current_user)):
    scope = user_scope_projects(user)
    unit_q: dict = {}
    exp_q: dict = {}
    stock_q: dict = {}
    if project_id:
        unit_q["project_id"] = project_id
        exp_q["project_id"] = project_id
        stock_q["project_id"] = project_id
    elif scope is not None:
        unit_q["project_id"] = {"$in": scope}
        exp_q["project_id"] = {"$in": scope}
        stock_q["project_id"] = {"$in": scope}

    SOLD_STATES = ["sale_confirmed", "post_sales_active", "fully_paid",
                   "registration_pending", "registered",
                   "possession_pending", "possession_completed",
                   "booked_pending_sales_approval"]
    total = await db.units.count_documents(unit_q)
    sold = await db.units.count_documents(
        {**unit_q, "status": {"$in": SOLD_STATES}})
    available = await db.units.count_documents(
        {**unit_q, "status": "available"})
    reserved = 0  # legacy — reservation removed

    units = await db.units.find(unit_q, {"_id": 0}).to_list(5000)
    accrued = sum(u.get("total_price", 0) or u.get("price", 0)
                  for u in units if u["status"] in SOLD_STATES)
    unit_ids = [u["unit_id"] for u in units]
    payments = await db.payments.find(
        {"unit_id": {"$in": unit_ids}}, {"_id": 0}
    ).to_list(10000)
    received = sum(p["amount"] for p in payments)

    projects_q: dict = {}
    if project_id:
        projects_q["project_id"] = project_id
    elif scope is not None:
        projects_q["project_id"] = {"$in": scope}
    projects = await db.projects.find(projects_q, {"_id": 0}).to_list(500)

    pending = await db.expenses.count_documents({**exp_q, "status": "pending"})
    stage1 = await db.expenses.count_documents(
        {**exp_q, "status": "stage1_approved"})
    approved = await db.expenses.count_documents(
        {**exp_q, "status": "final_approved"})
    rejected = await db.expenses.count_documents(
        {**exp_q, "status": "rejected"})

    tickets_open = await db.tickets.count_documents(
        {**({"project_id": exp_q["project_id"]} if exp_q else {}),
         "status": {"$in": ["open", "in_progress"]}})

    approved_amt = 0.0
    from collections import defaultdict
    by_day: dict[str, float] = defaultdict(float)
    vendor_now: dict[str, float] = defaultdict(float)
    vendor_prev: dict[str, float] = defaultdict(float)
    today = datetime.now(timezone.utc)
    this_month = today.strftime("%Y-%m")
    prev_dt = today.replace(day=1) - timedelta(days=1)
    prev_month = prev_dt.strftime("%Y-%m")

    async for e in db.expenses.find({**exp_q, "status": "final_approved"},
                                    {"_id": 0}):
        approved_amt += e["amount"]
        d = (e.get("final_at") or e["created_at"])[:10]
        by_day[d] += e["amount"]
        mo = d[:7]
        vendor = (e.get("vendor") or "Unknown").strip() or "Unknown"
        if mo == this_month:
            vendor_now[vendor] += e["amount"]
        elif mo == prev_month:
            vendor_prev[vendor] += e["amount"]

    trend = [{"date": d, "amount": v} for d, v in sorted(by_day.items())][-30:]

    vendors: list[dict] = []
    for v, amt in sorted(vendor_now.items(), key=lambda x: -x[1])[:5]:
        prev = vendor_prev.get(v, 0.0)
        delta_pct = None if prev == 0 else round(((amt - prev) / prev) * 100, 1)
        vendors.append({
            "vendor": v, "this_month": amt, "last_month": prev,
            "delta_pct": delta_pct,
        })

    # Per-project rollup
    per_project = []
    for p in projects:
        pu = [u for u in units if u["project_id"] == p["project_id"]]
        p_sold = sum(1 for u in pu if u["status"] in SOLD_STATES)
        p_accr = sum(u.get("total_price", 0) or u.get("price", 0)
                     for u in pu if u["status"] in SOLD_STATES)
        p_recv = sum(p2["amount"] for p2 in payments
                     if p2["project_id"] == p["project_id"])
        per_project.append({
            "project_id": p["project_id"],
            "name": p["name"],
            "project_type": p["project_type"],
            "location": p.get("location", ""),
            "units_total": len(pu),
            "units_sold": p_sold,
            "accrued": p_accr,
            "received": p_recv,
            "receivable": p_accr - p_recv,
        })

    return {
        "units": {"total": total, "sold": sold, "available": available,
                  "reserved": reserved},
        "revenue": {"accrued": accrued, "received": received,
                    "receivable": accrued - received},
        "expenses": {"pending": pending, "stage1": stage1,
                     "approved": approved, "rejected": rejected,
                     "approved_amount": approved_amt},
        "expense_trend": trend,
        "top_vendors": vendors,
        "tickets_open": tickets_open,
        "projects_count": len(projects),
        "per_project": per_project,
    }


async def _current_period_targets(project_ids: list[str]) -> dict:
    """Build monthly + quarterly current-period target vs received vs accrued."""
    keys = current_period_keys()
    out: dict[str, dict] = {}
    for kind in ("monthly", "quarterly"):
        k = keys[kind]
        tgt = (await _sum_targets(project_ids, kind, [k]))[k]
        act = (await _compute_period_actuals(project_ids, kind, [k]))[k]
        rec, acc = act["received"], act["accrued"]
        out[kind] = {
            "period_key": k,
            "target": tgt,
            "received": rec,
            "accrued": acc,
            "variance_received": rec - tgt,
            "variance_accrued": acc - tgt,
            "variance_received_pct": None if tgt == 0 else round(((rec - tgt) / tgt) * 100, 1),
            "variance_accrued_pct": None if tgt == 0 else round(((acc - tgt) / tgt) * 100, 1),
        }
    return out


# ------------------------------------------------------ search ------------
@api.get("/search")
async def global_search(q: str, user: User = Depends(get_current_user)):
    if not q:
        return {"projects": [], "units": [], "expenses": []}
    scope = user_scope_projects(user)
    proj_q = {"name": {"$regex": q, "$options": "i"}}
    if scope is not None:
        proj_q["project_id"] = {"$in": scope}
    projects = await db.projects.find(proj_q, {"_id": 0}).limit(10).to_list(10)

    unit_q = {"plot_number": {"$regex": q, "$options": "i"}}
    if scope is not None:
        unit_q["project_id"] = {"$in": scope}
    units = await db.units.find(unit_q, {"_id": 0}).limit(10).to_list(10)

    exp_q = {"$or": [{"category": {"$regex": q, "$options": "i"}},
                     {"vendor": {"$regex": q, "$options": "i"}}]}
    if scope is not None:
        exp_q["project_id"] = {"$in": scope}
    expenses = await db.expenses.find(exp_q, {"_id": 0}).limit(10).to_list(10)
    return {"projects": projects, "units": units, "expenses": expenses}


# ------------------------------------------------------ excel -------------
IMPORT_SHEETS = {
    "units": ["plot_number", "size", "facing", "price"],
    "stock_items": ["project_id", "name", "unit", "opening", "vendor"],
}


@api.get("/excel/template/{kind}")
async def excel_template(kind: str,
                         user: User = Depends(require_roles("super_admin", "process_admin"))):
    if kind not in IMPORT_SHEETS:
        raise HTTPException(400, "Unknown template kind")
    wb = Workbook()
    ws = wb.active
    ws.title = kind
    ws.append(IMPORT_SHEETS[kind])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="template_{kind}.xlsx"'})


@api.post("/excel/import/{kind}")
async def excel_import(kind: str, file: UploadFile = File(...),
                       user: User = Depends(require_roles("super_admin", "process_admin"))):
    if kind not in IMPORT_SHEETS:
        raise HTTPException(400, "Unknown import kind")
    wb = load_workbook(io.BytesIO(await file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty workbook")
    header = [str(c).strip() if c else "" for c in rows[0]]
    expected = IMPORT_SHEETS[kind]
    if header[:len(expected)] != expected:
        raise HTTPException(
            400, f"Header mismatch. Expected: {expected}")

    inserted = 0
    errors: list[dict] = []
    for i, row in enumerate(rows[1:], start=2):
        try:
            data = dict(zip(expected, row))
            if kind == "stock_items":
                pid = str(data.get("project_id") or "")
                if not pid or not await db.projects.find_one({"project_id": pid}):
                    raise ValueError("invalid project_id")
                it = StockItem(
                    project_id=pid,
                    name=str(data.get("name") or ""),
                    unit=str(data.get("unit") or "pcs"),
                    opening=float(data.get("opening") or 0),
                    vendor=str(data.get("vendor") or ""),
                )
                await db.stock_items.insert_one(it.model_dump())
            inserted += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    await audit(user.user_id, "excel_import", kind, "-",
                {"inserted": inserted, "errors": len(errors)})
    return {"inserted": inserted, "errors": errors}


# ------------------------------------------------------ excel export ------
def _rows_to_xlsx(sheet_name: str, headers: list[str],
                  rows: list[list[Any]]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0)
                      for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@api.get("/exports/units")
async def export_units(project_id: Optional[str] = None,
                       user: User = Depends(get_current_user)):
    q: dict = {}
    scope = user_scope_projects(user)
    if project_id:
        q["project_id"] = project_id
    elif scope is not None:
        q["project_id"] = {"$in": scope}
    units = await db.units.find(q, {"_id": 0}).to_list(5000)
    proj = {p["project_id"]: p["name"] for p in
            await db.projects.find({}, {"_id": 0}).to_list(1000)}
    headers = ["Plot #", "Project", "Size", "Facing", "Price",
               "Status", "Owner", "Contact", "Sold At"]
    rows = [[u["plot_number"], proj.get(u["project_id"], ""),
             u.get("size", ""), u.get("facing", ""), u.get("price", 0),
             u["status"], u.get("owner_name") or "",
             u.get("owner_contact") or "", u.get("sold_at") or ""]
            for u in units]
    buf = _rows_to_xlsx("units", headers, rows)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="units.xlsx"'})


@api.get("/exports/expenses")
async def export_expenses(project_id: Optional[str] = None,
                          user: User = Depends(get_current_user)):
    q: dict = {}
    scope = user_scope_projects(user)
    if project_id:
        q["project_id"] = project_id
    elif scope is not None:
        q["project_id"] = {"$in": scope}
    exps = await db.expenses.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    proj = {p["project_id"]: p["name"] for p in
            await db.projects.find({}, {"_id": 0}).to_list(1000)}
    users_map = {u["user_id"]: u["name"] for u in
                 await db.users.find({}, {"_id": 0}).to_list(1000)}
    headers = ["Date", "Project", "Category", "Vendor", "Amount",
               "Status", "Raised By", "Stage-1 By", "Final By",
               "Rejection Reason", "Description"]
    rows = [[e["created_at"][:10], proj.get(e["project_id"], ""),
             e["category"], e.get("vendor") or "", e["amount"], e["status"],
             users_map.get(e.get("raised_by"), ""),
             users_map.get(e.get("stage1_by"), ""),
             users_map.get(e.get("final_by"), ""),
             e.get("rejection_reason") or "", e.get("description") or ""]
            for e in exps]
    buf = _rows_to_xlsx("expenses", headers, rows)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="expenses.xlsx"'})


@api.get("/exports/payments")
async def export_payments(project_id: Optional[str] = None,
                          user: User = Depends(get_current_user)):
    q: dict = {}
    scope = user_scope_projects(user)
    if project_id:
        q["project_id"] = project_id
    elif scope is not None:
        q["project_id"] = {"$in": scope}
    pays = await db.payments.find(q, {"_id": 0}).sort("paid_on", -1).to_list(5000)
    units_map = {u["unit_id"]: u["plot_number"] for u in
                 await db.units.find({}, {"_id": 0}).to_list(5000)}
    proj = {p["project_id"]: p["name"] for p in
            await db.projects.find({}, {"_id": 0}).to_list(1000)}
    headers = ["Date", "Project", "Unit", "Amount", "Mode", "Reference"]
    rows = [[p["paid_on"], proj.get(p["project_id"], ""),
             units_map.get(p["unit_id"], ""), p["amount"], p["mode"],
             p.get("reference") or ""] for p in pays]
    buf = _rows_to_xlsx("payments", headers, rows)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="payments.xlsx"'})


@api.get("/exports/stock")
async def export_stock(project_id: Optional[str] = None,
                       user: User = Depends(get_current_user)):
    q: dict = {}
    scope = user_scope_projects(user)
    if project_id:
        q["project_id"] = project_id
    elif scope is not None:
        q["project_id"] = {"$in": scope}
    items = await db.stock_items.find(q, {"_id": 0}).to_list(5000)
    proj = {p["project_id"]: p["name"] for p in
            await db.projects.find({}, {"_id": 0}).to_list(1000)}
    headers = ["Item", "Project", "Unit", "Opening", "Inward", "Outward",
               "Closing", "Vendor"]
    rows = []
    for it in items:
        closing = it.get("opening", 0) + it.get("inward", 0) - it.get("outward", 0)
        rows.append([it["name"], proj.get(it["project_id"], ""), it["unit"],
                     it.get("opening", 0), it.get("inward", 0),
                     it.get("outward", 0), closing, it.get("vendor") or ""])
    buf = _rows_to_xlsx("stock", headers, rows)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="stock.xlsx"'})


# ------------------------------------------------------ onboarding -------
@api.get("/onboarding/status")
async def onboarding_status(user: User = Depends(get_current_user)):
    proj_count = await db.projects.count_documents({})
    acc_count = await db.users.count_documents({"role": {"$in": ["accounts_head", "accounts_rep"]}, "is_active": True})
    mgmt_count = await db.users.count_documents({"role": {"$in": ["process_admin", "crm_head"]}, "is_active": True})
    sales_count = await db.users.count_documents({"role": {"$in": ["sales_head", "sales_rep"]}, "is_active": True})
    crm_count = await db.users.count_documents({"role": {"$in": ["crm_head", "post_sales_rep"]}, "is_active": True})
    sm_count = await db.users.count_documents({"role": "site_supervisor", "is_active": True})
    units_count = await db.units.count_documents({})
    # Project has site_manager assigned
    proj_with_sm = await db.projects.count_documents(
        {"site_manager_id": {"$ne": None}})

    steps = [
        {"key": "password_reset", "label": "Change temporary password",
         "done": not user.must_reset_password},
        {"key": "add_management", "label": "Invite Management user",
         "done": mgmt_count > 0},
        {"key": "add_accounts", "label": "Invite Accounts user",
         "done": acc_count > 0},
        {"key": "add_sales", "label": "Invite Sales user",
         "done": sales_count > 0},
        {"key": "add_crm", "label": "Invite CRM user",
         "done": crm_count > 0},
        {"key": "add_project", "label": "Create your first project",
         "done": proj_count > 0},
        {"key": "assign_site_manager",
         "label": "Assign a Site Manager to a project",
         "done": proj_with_sm > 0 and sm_count > 0},
    ]
    done_count = sum(1 for s in steps if s["done"])
    return {
        "steps": steps,
        "done_count": done_count,
        "total_steps": len(steps),
        "system_ready": done_count == len(steps),
        "onboarding_completed": user.onboarding_completed,
        "counts": {"projects": proj_count, "units": units_count,
                   "accounts": acc_count, "management": mgmt_count,
                   "sales": sales_count, "crm": crm_count,
                   "site_supervisor": sm_count},
    }


@api.post("/onboarding/complete")
async def onboarding_complete(user: User = Depends(get_current_user)):
    await db.users.update_one({"user_id": user.user_id},
                              {"$set": {"onboarding_completed": True}})
    return {"ok": True}


# ------------------------------------------------------ dashboard config --
@api.get("/me/dashboard-config")
async def get_dash_config(user: User = Depends(get_current_user)):
    return {"widgets": user.dashboard_config.get("widgets", []) if user.dashboard_config else []}


@api.patch("/me/dashboard-config")
async def set_dash_config(payload: DashboardConfig,
                          user: User = Depends(get_current_user)):
    await db.users.update_one({"user_id": user.user_id},
                              {"$set": {"dashboard_config": payload.model_dump()}})
    return {"ok": True}


# ------------------------------------------------------ bulk unit import --
@api.post("/units/bulk-import")
async def units_bulk_import(project_id: str = Form(...),
                            file: UploadFile = File(...),
                            user: User = Depends(require_roles("super_admin", "process_admin"))):
    """Import units for a project via .xlsx or .csv."""
    proj = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not proj:
        raise HTTPException(404, "Project not found")

    raw = await file.read()
    rows: list[list[Any]] = []
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [row for row in reader]
    else:
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]

    if not rows:
        raise HTTPException(400, "Empty file")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_map = {h: i for i, h in enumerate(header)}
    if "plot_number" not in header_map:
        raise HTTPException(
            400,
            "Missing required column: plot_number. "
            "Expected columns: plot_number, size, facing, price")

    inserted, errors = 0, []
    for i, row in enumerate(rows[1:], start=2):
        try:
            def cell(name):
                idx = header_map.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            plot = cell("plot_number")
            if not plot:
                raise ValueError("plot_number required")
            price = cell("price") or 0
            try:
                price = float(price)
            except (TypeError, ValueError):
                price = 0.0
            u = Unit(project_id=project_id,
                     plot_number=str(plot).strip(),
                     size=str(cell("size") or "").strip(),
                     facing=str(cell("facing") or "").strip(),
                     price=price)
            await db.units.insert_one(u.model_dump())
            inserted += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    await audit(user.user_id, "units_bulk_import", "project", project_id,
                {"inserted": inserted, "errors": len(errors)})
    return {"inserted": inserted, "errors": errors}


@api.get("/units/bulk-template")
async def units_bulk_template(
        user: User = Depends(require_roles("super_admin", "process_admin"))):
    headers = ["plot_number", "size", "facing", "price"]
    wb = Workbook()
    ws = wb.active
    ws.title = "units"
    ws.append(headers)
    ws.append(["P-101", "1200 sqft", "North-East", 5000000])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="units_template.xlsx"'})


# ------------------------------------------------------ mount -------------
app.include_router(api)


@app.get("/health")
async def health_root():
    """Kubernetes liveness/readiness probe."""
    return {"status": "ok"}


@app.get("/api/health")
async def health_api():
    """Same as /health, prefixed for the app ingress."""
    return {"status": "ok"}


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)


@app.on_event("startup")
async def startup():
    init_storage()
    # seed admin if the env-configured one is missing
    existing = await db.users.find_one({"email": ADMIN_EMAIL}, {"_id": 0})
    if not existing:
        admin = User(
            email=ADMIN_EMAIL, name=ADMIN_NAME,
            role="super_admin", project_ids=[],
            password_hash=hash_pw(ADMIN_TEMP_PASSWORD),
            must_reset_password=True,
            is_active=True,
        )
        await db.users.insert_one(admin.model_dump())
        log.info("Seeded admin user: %s", ADMIN_EMAIL)
    else:
        # If the seeded admin is inactive (e.g. accidentally deactivated),
        # re-activate them so the app is never locked out entirely.
        if not existing.get("is_active", True):
            await db.users.update_one(
                {"email": ADMIN_EMAIL},
                {"$set": {"is_active": True}})
            log.info("Re-activated admin user: %s", ADMIN_EMAIL)
        # Always clear any stale lockout for the configured admin on startup.
        await db.login_attempts.delete_one({"_id": f"login:{ADMIN_EMAIL}"})
    # indexes
    try:
        await db.users.create_index("email", unique=True)
    except Exception:
        pass
    log.info("Backend started")


@app.on_event("shutdown")
async def shutdown():
    client.close()
