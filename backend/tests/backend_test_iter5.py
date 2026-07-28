"""
Iteration 5 backend tests - Password auth (JWT+bcrypt), onboarding, project
types, bulk unit import, dashboard config, user invite flow.
"""
import io
import os
import csv
import uuid
import time
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL",
                          "https://property-ops-60.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@vistaestates.com"
ADMIN_PW = "Vista@Admin#2026"

# --- shared admin session (unauth login uses the still-must-reset admin) ---
@pytest.fixture(scope="session")
def admin_token():
    """Log in as seed admin. If we've already reset the admin in a prior run,
    fall back to the changed password (but we prefer to keep admin pristine
    and instead use a brand new admin-created user for change-password test)."""
    r = requests.post(f"{API}/auth/login",
                      json={"email": ADMIN_EMAIL, "password": ADMIN_PW})
    if r.status_code == 200:
        return r.json()["access_token"]
    # Admin password may have been rotated in a prior test run — try known rotation
    r2 = requests.post(f"{API}/auth/login",
                       json={"email": ADMIN_EMAIL,
                             "password": os.environ.get(
                                 "ADMIN_ROTATED_PW", "Vista@Admin#2026")})
    if r2.status_code == 200:
        return r2.json()["access_token"]
    pytest.fail(f"Admin login failed: {r.status_code} {r.text}")


@pytest.fixture(scope="session")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


def _uniq(prefix="test5"):
    return f"{prefix}_{uuid.uuid4().hex[:8]}@example.com"


# ============================================================ AUTH TESTS ===
class TestAuth:
    def test_login_admin_success(self):
        r = requests.post(f"{API}/auth/login",
                          json={"email": ADMIN_EMAIL, "password": ADMIN_PW})
        assert r.status_code == 200, r.text
        data = r.json()
        assert "access_token" in data
        assert data["user"]["email"] == ADMIN_EMAIL
        assert "password_hash" not in data["user"]
        # must_reset_password may be True initially
        assert "must_reset_password" in data
        # cookie set
        assert "access_token" in r.cookies

    def test_auth_me_with_bearer(self, admin_headers):
        r = requests.get(f"{API}/auth/me", headers=admin_headers)
        assert r.status_code == 200
        d = r.json()
        assert d["email"] == ADMIN_EMAIL
        assert d.get("role") == "admin"

    def test_auth_me_no_token(self):
        r = requests.get(f"{API}/auth/me")
        assert r.status_code == 401

    def test_brute_force_lockout(self):
        # Use a unique email so we don't affect other tests
        email = _uniq("lockout")
        # Ensure not previously locked
        for _ in range(5):
            r = requests.post(f"{API}/auth/login",
                              json={"email": email, "password": "WrongPw!23"})
            assert r.status_code in (401, 429)
        r6 = requests.post(f"{API}/auth/login",
                           json={"email": email, "password": "WrongPw!23"})
        assert r6.status_code == 429, f"Expected lockout 429, got {r6.status_code}: {r6.text}"


# ================================================ change-password (new user) ===
class TestChangePasswordFlow:
    """Create a new user via admin, log in with temp pw, change password."""

    def test_create_user_and_change_password(self, admin_headers):
        email = _uniq("chpw")
        r = requests.post(f"{API}/users", headers=admin_headers,
                          json={"email": email, "name": "PW Tester",
                                "role": "management", "project_ids": []})
        assert r.status_code == 200, r.text
        payload = r.json()
        assert "user" in payload
        assert "temp_password" in payload
        assert "login_url" in payload
        assert payload["email_sent"] is False  # SMTP not configured
        assert "password_hash" not in payload["user"]
        temp_pw = payload["temp_password"]

        # login with temp pw
        r2 = requests.post(f"{API}/auth/login",
                           json={"email": email, "password": temp_pw})
        assert r2.status_code == 200, r2.text
        d2 = r2.json()
        assert d2["must_reset_password"] is True
        token = d2["access_token"]
        hdrs = {"Authorization": f"Bearer {token}"}

        # weak new password rejected
        r3 = requests.post(f"{API}/auth/change-password", headers=hdrs,
                           json={"current_password": temp_pw,
                                 "new_password": "abc"})
        assert r3.status_code == 400
        assert "password" in r3.text.lower()

        # valid change
        new_pw = "NewStrong#Pass2026"
        r4 = requests.post(f"{API}/auth/change-password", headers=hdrs,
                           json={"current_password": temp_pw,
                                 "new_password": new_pw})
        assert r4.status_code == 200, r4.text

        # old password fails
        r5 = requests.post(f"{API}/auth/login",
                           json={"email": email, "password": temp_pw})
        assert r5.status_code == 401

        # new pw works and must_reset_password = false
        r6 = requests.post(f"{API}/auth/login",
                           json={"email": email, "password": new_pw})
        assert r6.status_code == 200
        assert r6.json()["must_reset_password"] is False


# ============================================================ ONBOARDING ===
class TestOnboarding:
    def test_status_shape(self, admin_headers):
        r = requests.get(f"{API}/onboarding/status", headers=admin_headers)
        assert r.status_code == 200
        d = r.json()
        assert "steps" in d and "system_ready" in d
        assert "counts" in d
        for k in ["has_projects", "has_units", "has_accounts",
                  "has_management", "has_site_manager"]:
            assert k in d["steps"]
        # system_ready must equal AND of all steps
        assert d["system_ready"] == all(d["steps"].values())

    def test_complete_marks_user(self, admin_headers):
        r = requests.post(f"{API}/onboarding/complete", headers=admin_headers)
        assert r.status_code == 200
        m = requests.get(f"{API}/auth/me", headers=admin_headers).json()
        assert m.get("onboarding_completed") is True


# ============================================================ PROJECT TYPES ===
class TestProjectTypes:
    def test_types_returned(self):
        r = requests.get(f"{API}/projects/types")
        assert r.status_code == 200
        d = r.json()
        for t in ["residential", "commercial", "plot", "villa", "mixed"]:
            assert t in d
            assert "label" in d[t] and "unit_types" in d[t] and "fields" in d[t]

    def test_create_plot_with_extended_fields(self, admin_headers):
        payload = {
            "name": f"TEST_Plot_{uuid.uuid4().hex[:6]}",
            "project_type": "plot",
            "developer": "Acme Dev",
            "city": "Bangalore",
            "state": "KA",
            "pincode": "560001",
            "rera_number": "PRM/KA/RERA/1250/303/PR/2026/001",
            "start_date": "2026-01-15",
            "expected_completion": "2027-06-30",
            "total_units_planned": 120,
        }
        r = requests.post(f"{API}/projects", headers=admin_headers, json=payload)
        assert r.status_code == 200, r.text
        p = r.json()
        pid = p["project_id"]
        for k, v in payload.items():
            assert p[k] == v, f"{k}: {p[k]} != {v}"

        # PATCH update; other projects intact
        r2 = requests.patch(f"{API}/projects/{pid}", headers=admin_headers,
                            json={**payload, "developer": "New Dev",
                                  "city": "Mumbai"})
        assert r2.status_code == 200
        p2 = r2.json()
        assert p2["developer"] == "New Dev"
        assert p2["city"] == "Mumbai"
        # other project intact — check via list
        others = requests.get(f"{API}/projects", headers=admin_headers).json()
        assert any(x["project_id"] == pid for x in others)


# ============================================================ BULK IMPORT ===
@pytest.fixture(scope="class")
def residential_project(admin_headers):
    r = requests.post(f"{API}/projects", headers=admin_headers,
                      json={"name": f"TEST_Res_{uuid.uuid4().hex[:6]}",
                            "project_type": "residential",
                            "developer": "TestDev", "city": "Pune"})
    assert r.status_code == 200
    return r.json()["project_id"]


class TestBulkUnits:
    def test_template_xlsx(self, admin_headers):
        r = requests.get(
            f"{API}/units/bulk-template?project_type=residential",
            headers=admin_headers)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for h in ["unit_number", "unit_type", "price", "tower", "floor",
                  "bhk", "carpet_area_sqft"]:
            assert h in headers, f"missing header {h}"

    def test_bulk_import_xlsx(self, admin_headers, residential_project):
        # Build xlsx with 3 valid rows
        wb = Workbook()
        ws = wb.active
        headers = ["unit_number", "unit_type", "price", "tower", "floor",
                   "bhk", "carpet_area_sqft"]
        ws.append(headers)
        ws.append(["A-101", "2BHK", 5000000, "A", 1, "2BHK", 900])
        ws.append(["A-102", "2BHK", 5100000, "A", 1, "2BHK", 905])
        ws.append(["A-201", "3BHK", 7000000, "A", 2, "3BHK", 1200])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("units.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"project_id": residential_project}
        r = requests.post(f"{API}/units/bulk-import", headers=admin_headers,
                          data=data, files=files)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["inserted"] == 3
        assert d["errors"] == []

        # Verify persistence + attributes populated
        units = requests.get(f"{API}/units?project_id={residential_project}",
                             headers=admin_headers).json()
        assert len(units) >= 3
        a101 = next(u for u in units if u["unit_number"] == "A-101")
        assert a101["attributes"].get("tower") == "A"
        assert a101["attributes"].get("bhk") == "2BHK"
        # number coercion
        assert float(a101["attributes"].get("carpet_area_sqft")) == 900

    def test_bulk_import_csv(self, admin_headers, residential_project):
        csv_text = ("unit_number,unit_type,price,tower,floor,bhk\n"
                    "B-101,1BHK,3000000,B,1,1BHK\n"
                    "B-102,1BHK,3050000,B,1,1BHK\n")
        files = {"file": ("units.csv", csv_text.encode("utf-8"), "text/csv")}
        data = {"project_id": residential_project}
        r = requests.post(f"{API}/units/bulk-import", headers=admin_headers,
                          data=data, files=files)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["inserted"] == 2
        assert d["errors"] == []

    def test_bulk_import_missing_unit_number_error(self, admin_headers,
                                                   residential_project):
        # Row 2 has no unit_number, row 3 does
        csv_text = ("unit_number,unit_type,price\n"
                    ",1BHK,3000000\n"
                    "C-101,1BHK,3000000\n")
        files = {"file": ("units.csv", csv_text.encode("utf-8"), "text/csv")}
        r = requests.post(f"{API}/units/bulk-import", headers=admin_headers,
                          data={"project_id": residential_project},
                          files=files)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["inserted"] == 1
        assert len(d["errors"]) == 1
        assert d["errors"][0]["row"] == 2


# ============================================================ USERS ===
class TestUsersEndpoint:
    def test_password_hash_never_returned(self, admin_headers):
        r = requests.get(f"{API}/users", headers=admin_headers)
        assert r.status_code == 200
        for u in r.json():
            assert "password_hash" not in u

    def test_patch_user(self, admin_headers):
        email = _uniq("upd")
        c = requests.post(f"{API}/users", headers=admin_headers,
                          json={"email": email, "name": "Old",
                                "role": "accounts"}).json()
        uid = c["user"]["user_id"]
        r = requests.patch(f"{API}/users/{uid}", headers=admin_headers,
                           json={"name": "New Name"})
        assert r.status_code == 200
        assert r.json()["name"] == "New Name"
        assert "password_hash" not in r.json()


# ============================================================ DASHBOARD CONFIG ===
class TestDashboardConfig:
    def test_get_and_patch(self, admin_headers):
        r = requests.get(f"{API}/me/dashboard-config", headers=admin_headers)
        assert r.status_code == 200
        assert "widgets" in r.json()
        widgets = ["kpis", "vendors", "trend"]
        r2 = requests.patch(f"{API}/me/dashboard-config", headers=admin_headers,
                            json={"widgets": widgets})
        assert r2.status_code == 200
        r3 = requests.get(f"{API}/me/dashboard-config", headers=admin_headers)
        assert r3.json()["widgets"] == widgets


# ============================================================ SCOPED LISTING ===
class TestSiteManagerScope:
    def test_site_manager_sees_only_own_projects(self, admin_headers):
        # Create a project + 1 unit
        p_own = requests.post(f"{API}/projects", headers=admin_headers,
                              json={"name": f"TEST_SMOwn_{uuid.uuid4().hex[:6]}",
                                    "project_type": "residential"}).json()
        p_other = requests.post(f"{API}/projects", headers=admin_headers,
                                json={"name": f"TEST_SMOther_{uuid.uuid4().hex[:6]}",
                                      "project_type": "residential"}).json()
        # Unit in each
        requests.post(f"{API}/units", headers=admin_headers,
                      json={"project_id": p_own["project_id"],
                            "unit_number": "OWN-1", "price": 1000}).json()
        requests.post(f"{API}/units", headers=admin_headers,
                      json={"project_id": p_other["project_id"],
                            "unit_number": "OTH-1", "price": 1000}).json()

        # Create site_manager scoped to p_own
        email = _uniq("sm")
        created = requests.post(f"{API}/users", headers=admin_headers,
                                json={"email": email, "name": "Site M",
                                      "role": "site_manager",
                                      "project_ids": [p_own["project_id"]]}).json()
        temp_pw = created["temp_password"]
        # login, change pw
        tok = requests.post(f"{API}/auth/login",
                            json={"email": email,
                                  "password": temp_pw}).json()["access_token"]
        hdrs = {"Authorization": f"Bearer {tok}"}
        new_pw = "SMStrong#Pass2026"
        r_cp = requests.post(f"{API}/auth/change-password", headers=hdrs,
                             json={"current_password": temp_pw,
                                   "new_password": new_pw})
        assert r_cp.status_code == 200

        units = requests.get(f"{API}/units", headers=hdrs).json()
        pids = {u["project_id"] for u in units}
        assert pids == {p_own["project_id"]}, f"leaked projects: {pids}"


# ============================================================ REGRESSION ===
@pytest.fixture(scope="class")
def regression_project(admin_headers):
    p = requests.post(f"{API}/projects", headers=admin_headers,
                      json={"name": f"TEST_Reg_{uuid.uuid4().hex[:6]}",
                            "project_type": "residential"}).json()
    return p["project_id"]


class TestRegression:
    def test_dashboard_summary(self, admin_headers):
        r = requests.get(f"{API}/dashboard/summary", headers=admin_headers)
        assert r.status_code == 200, r.text

    def test_revenue_targets(self, admin_headers, regression_project):
        r = requests.post(f"{API}/revenue-targets", headers=admin_headers,
                          json={"project_id": regression_project,
                                "period_type": "monthly",
                                "period_key": "2026-02",
                                "amount": 500000})
        assert r.status_code == 200, r.text
        r2 = requests.get(f"{API}/revenue-targets?project_id="
                          f"{regression_project}", headers=admin_headers)
        assert r2.status_code == 200
        assert any(t["period_key"] == "2026-02" for t in r2.json())

    def test_exports_units(self, admin_headers):
        r = requests.get(f"{API}/exports/units", headers=admin_headers)
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "").lower() or \
               "excel" in r.headers.get("content-type", "").lower() or \
               "openxml" in r.headers.get("content-type", "").lower()

    def test_expense_flow(self, admin_headers, regression_project):
        r = requests.post(f"{API}/expenses", headers=admin_headers,
                          json={"project_id": regression_project,
                                "category": "materials", "amount": 100000,
                                "vendor": "V", "description": "d"})
        assert r.status_code == 200, r.text
        eid = r.json()["expense_id"]
        r2 = requests.post(f"{API}/expenses/{eid}/stage1", headers=admin_headers,
                           json={"action": "approve"})
        assert r2.status_code == 200
        r3 = requests.post(f"{API}/expenses/{eid}/final", headers=admin_headers,
                           json={"action": "approve"})
        assert r3.status_code == 200
